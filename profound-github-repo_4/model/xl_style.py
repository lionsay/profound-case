"""
Workbook formatting standard.

Font convention (applied by construction, never by hand):
    dark blue   hardcoded input        (lives on Inputs tab ONLY)
    black       calculation on this sheet
    purple      reference to another cell on the SAME sheet
    dark green  link to ANOTHER sheet
    red         formula that deliberately breaks the row's pattern

Structure rules: no gridlines, nothing in row 1 or column A, no merged cells,
one row = one formula carried across all periods, single scenario switch.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- palette
BLUE = "0000C0"      # hardcoded input
BLACK = "000000"     # same-sheet calculation
PURPLE = "7030A0"    # same-sheet reference
GREEN = "006100"     # cross-sheet link
RED = "C00000"       # deliberate break in row pattern
GREY = "808080"      # annotation
NAVY = "1F3864"      # section headers

FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FILL_SUBHEAD = PatternFill("solid", fgColor="D9E2F3")
FILL_INPUT = PatternFill("solid", fgColor="FFF9E6")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_BAND = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
MED = Side(style="medium", color="1F3864")
BORDER_UNDER = Border(bottom=THIN)
BORDER_SECTION = Border(bottom=MED)

# --- number formats
FMT_M = '#,##0.0;(#,##0.0);"-"'          # $ millions, 1dp
FMT_K = '#,##0;(#,##0);"-"'              # whole numbers
FMT_0 = '#,##0;(#,##0);"-"'
FMT_1 = '#,##0.0;(#,##0.0);"-"'
FMT_2 = '#,##0.00;(#,##0.00);"-"'
FMT_PCT0 = '0%;(0%);"-"'
FMT_PCT1 = '0.0%;(0.0%);"-"'
FMT_X = '0.00"x";(0.00"x");"-"'
FMT_DATE = 'mmm-yy'
FMT_USD0 = '$#,##0;($#,##0);"-"'
FMT_USD_K = '$#,##0,"K";($#,##0,"K");"-"'

# --- geometry
COL_LABEL = 2      # B
COL_UNITS = 3      # C
COL_M0 = 4         # D  first month column
N_MONTHS = 60
COL_MLAST = COL_M0 + N_MONTHS - 1          # BK
COL_FY0 = COL_MLAST + 2                    # BM  first fiscal-year column
N_YEARS = 5
COL_FYLAST = COL_FY0 + N_YEARS - 1

TAB_COLORS = {
    "Cover": "1F3864", "Inputs": "C00000", "Rev-Capacity": "2E7D32",
    "Rev-Cohorts": "2E7D32", "Rev-ARR Bridge": "2E7D32", "Headcount": "2E7D32",
    "P&L": "0070C0", "Cash Flow": "0070C0", "KPIs": "7030A0",
    "Scenarios": "7030A0", "Checks": "BF8F00",
}


def col(i):
    return get_column_letter(i)


def setup_sheet(ws, title=None, freeze="D7"):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 9
    for i in range(COL_M0, COL_MLAST + 1):
        ws.column_dimensions[col(i)].width = 11.5
    ws.column_dimensions[col(COL_MLAST + 1)].width = 2.5
    for i in range(COL_FY0, COL_FYLAST + 1):
        ws.column_dimensions[col(i)].width = 13
    if ws.title in TAB_COLORS:
        ws.sheet_properties.tabColor = TAB_COLORS[ws.title]
    return ws


def title_block(ws, title, subtitle=""):
    c = ws.cell(row=2, column=COL_LABEL, value=title)
    c.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    if subtitle:
        s = ws.cell(row=3, column=COL_LABEL, value=subtitle)
        s.font = Font(name="Calibri", size=9, italic=True, color=GREY)


def section(ws, row, text, last_col=None):
    last_col = last_col or COL_FYLAST
    c = ws.cell(row=row, column=COL_LABEL, value=text)
    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    for i in range(COL_LABEL, last_col + 1):
        ws.cell(row=row, column=i).fill = FILL_HEADER
    return row + 1


def subsection(ws, row, text, last_col=None):
    last_col = last_col or COL_FYLAST
    c = ws.cell(row=row, column=COL_LABEL, value=text)
    c.font = Font(name="Calibri", size=9, bold=True, color=NAVY)
    for i in range(COL_LABEL, last_col + 1):
        ws.cell(row=row, column=i).border = BORDER_UNDER
    return row + 1


def label(ws, row, text, units="", indent=0, bold=False, italic=False):
    c = ws.cell(row=row, column=COL_LABEL, value=("    " * indent) + text)
    c.font = Font(name="Calibri", size=9.5, bold=bold, italic=italic)
    u = ws.cell(row=row, column=COL_UNITS, value=units)
    u.font = Font(name="Calibri", size=8, color=GREY)
    u.alignment = Alignment(horizontal="right")
    return row


def write_row(ws, row, formula_fn, fmt=FMT_M, color=BLACK, bold=False,
              start=COL_M0, end=None, band=False):
    """Write one formula carried across every period column. formula_fn(i, n)
    receives the absolute column index and the zero-based period number."""
    end = end or COL_MLAST
    for n, i in enumerate(range(start, end + 1)):
        v = formula_fn(i, n)
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name="Calibri", size=9.5, color=color, bold=bold)
        c.number_format = fmt
        if band:
            c.fill = FILL_BAND
    return row


def const_row(ws, row, values, fmt=FMT_M, color=BLUE, start=COL_M0, bold=False):
    for n, v in enumerate(values):
        c = ws.cell(row=row, column=start + n, value=v)
        c.font = Font(name="Calibri", size=9.5, color=color, bold=bold)
        c.number_format = fmt
        c.fill = FILL_INPUT if color == BLUE else PatternFill()
    return row


def note(ws, row, text):
    c = ws.cell(row=row, column=COL_LABEL, value=text)
    c.font = Font(name="Calibri", size=8, italic=True, color=GREY)
    return row + 1
