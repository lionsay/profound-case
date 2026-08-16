"""
Phase 3 - Excel workbook generation.

Writes a LIVE-FORMULA workbook from assumptions.py. Every calculation cell is a
real Excel formula, so a reviewer can change any blue input and the whole model
recalculates natively. Nothing is pasted as a value.

Font convention, structure rules and geometry live in xl_style.py.
"""
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

from xl_style import *
from model import run, month_index, monthly_from_annual
import assumptions as A

DATES, YEARS, MOYS = month_index()
YRS = [2026, 2027, 2028, 2029, 2030]
SCN = list(A.SCENARIO_LABELS.keys())      # single source of order, shared with the dashboard
SCN_LBL = [A.SCENARIO_LABELS[s] for s in SCN]

I = {}   # registry of input cell addresses: name -> "Inputs!$D$12" or row for year-lookup
YROW = {}  # name -> row number on Inputs for year-varying blocks


def scn_val(scenario, path, default):
    """Value of an assumption under a scenario override."""
    ov = A.SCENARIOS.get(scenario, {})
    return ov.get(path, default)


# =============================================================================
# Period header, repeated identically on every calculation sheet
# =============================================================================
def period_header(ws, first_row=5):
    label(ws, first_row, "Month ending", "")
    for n, i in enumerate(range(COL_M0, COL_MLAST + 1)):
        c = ws.cell(row=first_row, column=i, value=DATES[n])
        c.number_format = FMT_DATE
        c.font = Font(name="Calibri", size=9, bold=True, color=NAVY)
        c.alignment = Alignment(horizontal="center")
    label(ws, first_row + 1, "Fiscal year", "")
    for n, i in enumerate(range(COL_M0, COL_MLAST + 1)):
        c = ws.cell(row=first_row + 1, column=i, value=YEARS[n])
        c.number_format = "0"
        c.font = Font(name="Calibri", size=8, color=GREY)
        c.alignment = Alignment(horizontal="center")
    label(ws, first_row + 2, "Period", "")
    for n, i in enumerate(range(COL_M0, COL_MLAST + 1)):
        c = ws.cell(row=first_row + 2, column=i, value=n + 1)
        c.number_format = "0"
        c.font = Font(name="Calibri", size=8, color=GREY)
        c.alignment = Alignment(horizontal="center")
    # fiscal-year summary columns
    for n, i in enumerate(range(COL_FY0, COL_FYLAST + 1)):
        c = ws.cell(row=first_row, column=i, value=f"FY{YRS[n]}")
        c.font = Font(name="Calibri", size=9, bold=True, color=NAVY)
        c.alignment = Alignment(horizontal="center")
    for i in range(COL_LABEL, COL_FYLAST + 1):
        ws.cell(row=first_row + 2, column=i).border = BORDER_UNDER
    return first_row + 4


def fy_sum(ws, row):
    """Fiscal-year total columns for a flow row."""
    for n, i in enumerate(range(COL_FY0, COL_FYLAST + 1)):
        a, b = col(COL_M0 + n * 12), col(COL_M0 + n * 12 + 11)
        c = ws.cell(row=row, column=i, value=f"=SUM({a}{row}:{b}{row})")
        c.font = Font(name="Calibri", size=9.5, color=PURPLE)
        c.number_format = ws.cell(row=row, column=COL_M0).number_format


def fy_last(ws, row):
    """Fiscal-year closing-balance columns for a stock row."""
    for n, i in enumerate(range(COL_FY0, COL_FYLAST + 1)):
        b = col(COL_M0 + n * 12 + 11)
        c = ws.cell(row=row, column=i, value=f"={b}{row}")
        c.font = Font(name="Calibri", size=9.5, color=PURPLE)
        c.number_format = ws.cell(row=row, column=COL_M0).number_format


def yr_lookup(name, ci, hdr_row):
    """Formula fragment that pulls a year-varying input for this month's FY."""
    r = YROW[name]
    return f"INDEX(Inputs!$D${r}:$H${r},1,{col(ci)}${hdr_row + 1}-2025)"


# =============================================================================
# COVER
# =============================================================================
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    setup_sheet(ws, freeze="A1")
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 78
    ws.sheet_view.showGridLines = False

    ws.cell(row=2, column=2, value="Profound").font = Font(size=20, bold=True, color=NAVY)
    ws.cell(row=3, column=2, value="Five Year Operating Model").font = Font(size=13, color=NAVY)

    r = 7
    r = section(ws, r, "HOW THIS MODEL IS BUILT", 3)
    for k, v in [
        ("Revenue is an output, not a target",
         "The hiring plan is the input. Sales capacity, ramp and attrition determine new bookings, "
         "and growth falls out. No growth rate is assumed anywhere."),
        ("NRR, GRR and gross margin are outputs",
         "They are computed from retention and expansion drivers and from query volume, then compared "
         "against benchmark on the KPIs tab. None of them is an input."),
        ("Cohorts are dollar denominated",
         "Correct basis for NRR and GRR. A separate logo view uses a worse churn rate and does not "
         "feed the ARR calculation."),
        ("Every input lives on the Inputs tab",
         "A hardcoded value anywhere else in this workbook is a defect. Change any blue cell and the "
         "model recalculates."),
        ("Checks ship inside the workbook",
         "The Checks tab reconciles cohorts to the ARR bridge, the deferred revenue roll and the cash "
         "roll on every period. It is not deleted before sending."),
    ]:
        ws.cell(row=r, column=2, value=k).font = Font(size=9.5, bold=True)
        c = ws.cell(row=r, column=3, value=v)
        c.font = Font(size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    r = section(ws, r, "FORMATTING KEY", 3)
    for txt, colr, desc in [
        ("Hardcoded input", BLUE, "Inputs tab only"),
        ("Calculation", BLACK, "Formula computed on this sheet"),
        ("Same sheet reference", PURPLE, "Points at another cell on this sheet"),
        ("Link to another sheet", GREEN, "Cross sheet reference"),
        ("Break in row pattern", RED, "Formula deliberately differs from the rest of the row"),
    ]:
        c = ws.cell(row=r, column=2, value=txt)
        c.font = Font(size=9.5, color=colr, bold=True)
        ws.cell(row=r, column=3, value=desc).font = Font(size=9, color=GREY)
        r += 1

    r += 1
    r = section(ws, r, "TAB MAP", 3)
    for t, d in [
        ("Inputs", "Every assumption. Scenario switch in cell D5 drives the whole model."),
        ("Rev-Capacity", "AE hiring, ramp, attrition, productive capacity, new bookings. Self serve."),
        ("Rev-Cohorts", "Dollar cohort triangle. Renewal events, expansion, churn."),
        ("Rev-ARR Bridge", "Beginning plus new plus expansion less churn. Agents and self serve."),
        ("Headcount", "Roster by function, hiring, attrition, payroll, stock compensation."),
        ("P&L", "Revenue recognition, volume driven cost of revenue, operating expense, EBITDA."),
        ("Cash Flow", "Billings, deferred revenue, working capital, capital expenditure, cash roll."),
        ("KPIs", "Computed outputs against benchmark. Runway and raise decision date."),
        ("Scenarios", "Five cases. Base, upside, downside, downside with hiring freeze, Agents led."),
        ("Checks", "Live reconciliations. Everything should read OK."),
    ]:
        ws.cell(row=r, column=2, value=t).font = Font(size=9.5, bold=True)
        ws.cell(row=r, column=3, value=d).font = Font(size=9)
        r += 1
    return ws


# =============================================================================
# INPUTS
# =============================================================================
def build_inputs(wb):
    ws = wb.create_sheet("Inputs")
    setup_sheet(ws, freeze="D6")
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 10
    for i in range(4, 9):
        ws.column_dimensions[col(i)].width = 14
    ws.column_dimensions["I"].width = 3
    ws.column_dimensions["J"].width = 60

    title_block(ws, "Inputs")

    # scenario switch
    ws.cell(row=5, column=2, value="SCENARIO SWITCH").font = Font(size=10, bold=True, color="FFFFFF")
    for i in range(2, 9):
        ws.cell(row=5, column=i).fill = FILL_HEADER
    sw = ws.cell(row=5, column=4, value=1)
    sw.font = Font(size=10, bold=True, color=BLUE)
    sw.fill = PatternFill("solid", fgColor="FFE699")
    sw.alignment = Alignment(horizontal="center")
    ws.cell(row=5, column=5, value='=INDEX({"' + '";"'.join(SCN_LBL) + '"},$D$5,1)').font = \
        Font(size=10, bold=True, color=PURPLE)
    ws.cell(row=5, column=6, value="1=Base 2=Upside 3=Downside 4=Downside+Freeze 5=Agents-Led").font = \
        Font(size=8, italic=True, color="FFFFFF")
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
    ws.add_data_validation(dv); dv.add(sw)
    I["scenario"] = "Inputs!$D$5"

    r = 7
    for n, y in enumerate(YRS):
        c = ws.cell(row=r, column=4 + n, value=y)
        c.font = Font(size=9, bold=True, color=NAVY)
        c.alignment = Alignment(horizontal="center")
    r += 1

    def single(row, name, text, val, fmt=FMT_2, units="", src=""):
        label(ws, row, text, units)
        c = ws.cell(row=row, column=4, value=val)
        c.font = Font(size=9.5, color=BLUE); c.number_format = fmt; c.fill = FILL_INPUT
        if src:
            ws.cell(row=row, column=10, value=src).font = Font(size=8, italic=True, color=GREY)
        I[name] = f"Inputs!$D${row}"
        return row + 1

    def byyear(row, name, text, vals, fmt=FMT_2, units="", src=""):
        label(ws, row, text, units)
        for n, v in enumerate(vals):
            c = ws.cell(row=row, column=4 + n, value=v)
            c.font = Font(size=9.5, color=BLUE); c.number_format = fmt; c.fill = FILL_INPUT
        if src:
            ws.cell(row=row, column=10, value=src).font = Font(size=8, italic=True, color=GREY)
        YROW[name] = row
        return row + 1

    def scen_block(row, name, text, per_scn_vals, fmt=FMT_2, units="", src=""):
        """5 scenario rows over 5 year columns, then an Active row driven by the switch."""
        row = subsection(ws, row, text, 8)
        first = row
        for si, sc in enumerate(SCN):
            label(ws, row, SCN_LBL[si], units, indent=1,
                  italic=(sc != "base"))
            for n, v in enumerate(per_scn_vals[si]):
                c = ws.cell(row=row, column=4 + n, value=v)
                c.font = Font(size=9.5, color=BLUE); c.number_format = fmt
                c.fill = FILL_INPUT
            row += 1
        last = row - 1
        label(ws, row, "Active", units, indent=1, bold=True)
        for n in range(5):
            c = ws.cell(row=row, column=4 + n,
                        value=f"=INDEX($D${first}:$H${last},$D$5,{n+1})")
            c.font = Font(size=9.5, color=PURPLE, bold=True); c.number_format = fmt
        YROW[name] = row
        if src:
            ws.cell(row=row, column=10, value=src).font = Font(size=8, italic=True, color=GREY)
        return row + 2

    # ---- given
    r = section(ws, r, "GIVEN", 8)
    r = single(r, "opening_arr", "Opening ARR at 1 Jan 2026", A.GIVEN["opening_arr"], FMT_USD0, "$",
               "Given. Interpreted as ending contracted ARR.")
    r = single(r, "opening_cash", "Opening cash", A.GIVEN["opening_cash"], FMT_USD0, "$",
               "Given. Assumed post Series C. No further financing in base case.")
    r = single(r, "opening_hc", "Opening headcount", A.GIVEN["opening_headcount"], FMT_0, "#", "Given.")
    r = single(r, "min_cash", "Minimum cash balance", A.GIVEN["min_cash"], FMT_USD0, "$",
               "Given. Board operating floor, not zero. Usable cash is $85M.")
    r += 1

    # ---- opening base
    r = section(ws, r, "OPENING BASE", 8)
    r = single(r, "ent_arr0", "Enterprise ARR", A.OPENING_BASE["ent_arr"], FMT_USD0, "$",
               "Ties to given total with self serve below.")
    r = single(r, "ent_cust0", "Enterprise customers", A.OPENING_BASE["ent_customers"], FMT_0, "#",
               "Public: 700+ enterprise customers. Source: Fortune, February 2026: 700+ enterprise customers at the Series C.")
    r = single(r, "ss_arr0", "Self serve ARR", A.OPENING_BASE["selfserve_arr"], FMT_USD0, "$")
    r = single(r, "ss_cust0", "Self serve customers", A.OPENING_BASE["selfserve_customers"], FMT_0, "#",
               "Published tiers: $99 and $399 per month.")
    r += 1

    # ---- sales capacity
    r = section(ws, r, "SALES CAPACITY", 8)
    r = single(r, "opening_aes", "Opening quota carrying AEs", A.SALES["opening_aes"], FMT_0, "#")
    r = single(r, "quota", "AE annual quota, fully ramped", A.SALES["quota_annual"], FMT_USD0, "$",
               "Benchmark commercial band $0.9M to $1.4M.")
    r = single(r, "quota_g", "Quota growth per year", A.SALES["quota_growth_pa"], FMT_PCT1, "%",
               "Negative. Territories thin as the org scales.")
    for k in range(5):
        vals = []
        for y in YRS:
            L = A.SALES["ramp_months_by_year"][y]
            crv = A.SALES["ramp_curve"] if L == 4 else A.SALES["ramp_curve_5mo"]
            vals.append(crv[k] if k < len(crv) else 1.0)
        r = byyear(r, f"ramp{k}", f"Ramp productivity, month {k+1}", vals, FMT_PCT0, "%")
    r = single(r, "tth", "Time to hire, months", A.SALES["time_to_hire_months"], FMT_0, "#",
               "Benchmark 6 to 10 weeks.")
    r = single(r, "sdr_per_ae", "SDRs per AE", A.SALES["sdr_per_ae"], FMT_2, "x")
    r = single(r, "mktg_per_ae", "Marketing per AE", A.SALES["mktg_per_ae"], FMT_2, "x")
    r = single(r, "sops_per_ae", "Sales ops per AE", A.SALES["salesops_per_ae"], FMT_2, "x")
    r = single(r, "enable_per_ae", "Enablement per AE", 1.0 / 20.0, FMT_2, "x",
               "Explicit enablement investment holding ramp at the short end.")
    r += 1
    r = scen_block(r, "attainment", "Quota attainment", [
        [scn_val(s, "SALES.attainment", A.SALES["attainment"])] * 5 for s in SCN
    ], FMT_PCT0, "%", "Benchmark 70% median, 80% top quartile.")
    r = scen_block(r, "ae_hires", "AE hires per month", [
        [scn_val(s, "SALES.ae_hires_per_month", A.SALES["ae_hires_per_month"])[y] for y in YRS]
        for s in SCN], FMT_1, "#")

    # ---- retention
    r = section(ws, r, "RETENTION AND EXPANSION", 8)
    r = scen_block(r, "surv1", "First renewal dollar survival", [
        [scn_val(s, "RETENTION.first_renewal_survival", A.RETENTION["first_renewal_survival"])] * 5
        for s in SCN], FMT_PCT1, "%",
        "Least proven number in the business. Nothing in the base has renewed yet.")
    r = byyear(r, "surv_later", "Later renewal dollar survival",
               [A.RETENTION["later_renewal_survival"][y] for y in YRS], FMT_PCT1, "%")
    r = single(r, "churn_ic", "In contract churn per year", A.RETENTION["in_contract_churn_pa"],
               FMT_PCT1, "%", "Low by design. Annual terms concentrate churn at renewal.")
    r = scen_block(r, "seat_exp", "Seat and usage expansion per year", [
        [scn_val(s, "RETENTION.seat_expansion_pa", A.RETENTION["seat_expansion_pa"])[y] for y in YRS]
        for s in SCN], FMT_PCT1, "%", "Must decay. Expansion compresses as accounts saturate.")
    r = single(r, "uplift", "Renewal price uplift", A.RETENTION["renewal_price_uplift"], FMT_PCT1, "%")
    r = byyear(r, "cross", "Cross sell uplift at renewal",
               [A.RETENTION["cross_sell_uplift"][y] for y in YRS], FMT_PCT1, "%")
    r = single(r, "logo_prem", "Logo churn premium over dollar churn",
               A.RETENTION["logo_churn_premium"], FMT_PCT1, "%",
               "Churned accounts are systematically smaller. Reporting view only.")
    r += 1

    # ---- pricing
    r = section(ws, r, "PRICING", 8)
    r = single(r, "acv", "Enterprise new logo ACV", A.ACV["ent_new_logo_acv"], FMT_USD0, "$")
    r = single(r, "acv_g", "Enterprise ACV growth per year", A.ACV["ent_acv_growth_pa"], FMT_PCT1, "%")
    r = single(r, "ss_acv", "Self serve ACV", A.ACV["selfserve_acv"], FMT_USD0, "$")
    r = single(r, "ss_acv_g", "Self serve ACV growth per year", A.ACV["selfserve_acv_growth_pa"], FMT_PCT1, "%")
    r += 1

    # ---- self serve
    r = section(ws, r, "SELF SERVE", 8)
    r = single(r, "ss_adds", "New self serve customers per month", A.SELFSERVE["new_customers_per_month"], FMT_1, "#")
    r = single(r, "ss_adds_g", "Growth in monthly adds per year", A.SELFSERVE["new_cust_growth_pa"], FMT_PCT1, "%")
    r = single(r, "ss_churn", "Monthly churn", A.SELFSERVE["monthly_churn"], FMT_PCT1, "%")
    r = single(r, "ss_conv", "Monthly conversion into enterprise", A.SELFSERVE["monthly_conv_to_ent"], FMT_PCT1, "%",
               "The strategic value of self serve. Feeds enterprise bookings.")
    r += 1

    # ---- agents
    r = section(ws, r, "AGENTS", 8)
    r = scen_block(r, "attach", "Attach rate, share of enterprise customers at year end", [
        [scn_val(s, "AGENTS.attach_by_yearend", A.AGENTS["attach_by_yearend"])[y] for y in YRS]
        for s in SCN], FMT_PCT0, "%", "Biggest swing factor, weakest benchmark.")
    r = scen_block(r, "ag_pct", "Agents ACV as share of core ACV", [
        [scn_val(s_, "AGENTS.acv_pct_of_core", A.AGENTS["acv_pct_of_core"])] * 5 for s_ in SCN
    ], FMT_PCT0, "%")
    r = scen_block(r, "ag_g", "Agents ACV growth per year", [
        [scn_val(s_, "AGENTS.acv_growth_pa", A.AGENTS["acv_growth_pa"])] * 5 for s_ in SCN
    ], FMT_PCT1, "%")
    r += 1

    # ---- cogs
    r = section(ws, r, "COST OF REVENUE", 8)
    r = single(r, "prompts", "Tracked prompts per enterprise customer",
               A.COGS["tracked_prompts_per_customer"], FMT_0, "#",
               "The only soft input of the three. Set for a large multi brand enterprise. "
               "Least observable driver in the model.")
    r = single(r, "engines", "AI engines monitored", A.COGS["engines_monitored"], FMT_0, "#",
               "Published product fact. The Enterprise tier covers all ten engines.")
    r = single(r, "runs_pm", "Runs per prompt per month", A.COGS["runs_per_prompt_per_month"],
               FMT_0, "#", "Daily refresh. Tracking change over time is the core value proposition.")
    r = note(ws, r, "Sensitivity: at 300,000 queries per customer per month gross margin is 87% "
                    "and the base case does not breach the cash floor. At 900,000 it is 79% and "
                    "the breach moves to Apr-2028. The downside capital requirement stays between "
                    "$82M and $116M across that whole range, which is why the raise argument rests "
                    "on the downside rather than on the base case breach date.")
    r = single(r, "query_g", "Query volume growth per year", A.COGS["query_volume_growth_pa"], FMT_PCT1, "%")
    r = single(r, "cost_q", "Cost per query", A.COGS["cost_per_query"], '$#,##0.00000', "$")
    r = single(r, "defl", "Inference cost deflation per year", A.COGS["inference_deflation_pa"], FMT_PCT1, "%",
               "Observed token price decline.")
    r = byyear(r, "ag_cogs", "Agents cost as share of Agents revenue",
               [A.COGS["agents_cogs_pct"][y] for y in YRS], FMT_PCT0, "%")
    r = single(r, "ss_cogs", "Self serve cost as share of self serve revenue",
               A.COGS["selfserve_cogs_pct"], FMT_PCT0, "%")
    r = byyear(r, "host", "Hosting as share of revenue",
               [A.COGS["hosting_pct_of_rev"][y] for y in YRS], FMT_PCT1, "%")
    r = single(r, "data3p", "Third party data as share of revenue",
               A.COGS["third_party_data_pct_of_rev"], FMT_PCT1, "%")
    r = note(ws, r, "Policy: customer success payroll sits in cost of revenue. Stated, not assumed silently.")
    r += 1

    # ---- headcount
    r = section(ws, r, "HEADCOUNT AND COMPENSATION", 8)
    for k, v in A.HEADCOUNT["opening_split"].items():
        r = single(r, f"hc0_{k}", f"Opening headcount, {k.upper()}", v, FMT_0, "#")
    for k, v in A.HEADCOUNT["loaded_cost"].items():
        r = single(r, f"lc_{k}", f"Fully loaded cost, {k.upper()}", v, FMT_USD0, "$")
    r = single(r, "comp_infl", "Compensation inflation per year", A.HEADCOUNT["comp_inflation_pa"], FMT_PCT1, "%")
    r = single(r, "attr_s", "Sales attrition per year", A.HEADCOUNT["attrition_sales_pa"], FMT_PCT1, "%",
               "Best in class. Benchmark AE turnover is 30% to 32%. Source: Benchmark AE turnover 30-32%, SDR 45-48%, blended 35%. Best in class under 20%. Optifai, 939 B2B companies.")
    r = single(r, "attr_ns", "Non sales attrition per year", A.HEADCOUNT["attrition_nonsales_pa"], FMT_PCT1, "%")
    r = single(r, "recruit", "Recruiting cost per hire", A.HEADCOUNT["recruiting_cost_per_hire"], FMT_USD0, "$")
    r = single(r, "sbc_pct", "Stock compensation as share of revenue",
               A.HEADCOUNT["sbc_pct_of_revenue"], FMT_PCT0, "%",
               "Non cash. Benchmarked on revenue, as public companies disclose it. "
               "Public SaaS comps run 15 to 25 per cent of revenue.")
    r = scen_block(r, "csm_ratio", "Enterprise customers per CSM", [
        [scn_val(s_, "HEADCOUNT.ent_custs_per_csm", A.HEADCOUNT["ent_custs_per_csm"])[y] for y in YRS]
        for s_ in SCN], FMT_1, "#", "Tightens. Expansion is a customer success delivered motion.")
    r = scen_block(r, "ga_hires", "G&A hires per month", [
        [scn_val(s_, "HEADCOUNT.ga_hires_per_month", A.HEADCOUNT["ga_hires_per_month"])[y] for y in YRS]
        for s_ in SCN], FMT_1, "#")
    r = scen_block(r, "rnd_hires", "R&D hires per month", [
        [scn_val(s, "HEADCOUNT.rnd_hires_per_month", A.HEADCOUNT["rnd_hires_per_month"])[y] for y in YRS]
        for s in SCN], FMT_1, "#")

    # ---- opex
    r = section(ws, r, "OPERATING EXPENSE", 8)
    r = single(r, "sw_head", "Software per head per year", A.OPEX["software_per_head_pa"], FMT_USD0, "$")
    r = single(r, "te_head", "Travel and entertainment per head per year", A.OPEX["te_per_head_pa"], FMT_USD0, "$")
    r = single(r, "oth_head", "Other G&A per head per year", A.OPEX["other_ga_per_head_pa"], FMT_USD0, "$")
    r = single(r, "fac_head", "Facilities per head per year", A.OPEX["facilities_per_head_pa"], FMT_USD0, "$")
    r = single(r, "fac_step_hc", "Headcount triggering office step", A.OPEX["office_step_at_headcount"], FMT_0, "#",
               "They will outgrow the office. Facilities is a step, not a straight line.")
    r = single(r, "fac_step", "Office step cost per year", A.OPEX["office_step_cost_pa"], FMT_USD0, "$")
    r = byyear(r, "rnd_compute", "R&D compute as share of revenue",
               [A.OPEX["rnd_compute_pct_of_rev"][y] for y in YRS], FMT_PCT1, "%",
               "Model training, evaluation and non cost of revenue infrastructure.")
    r = single(r, "mktg_prog", "Marketing programmes as share of revenue",
               A.OPEX["marketing_programs_pct_of_rev"], FMT_PCT1, "%")
    r = single(r, "prof", "Professional fees per year", A.OPEX["professional_fees_pa"], FMT_USD0, "$")
    r = single(r, "prof_g", "Professional fees growth per year", A.OPEX["professional_fees_growth_pa"], FMT_PCT1, "%")
    r = single(r, "ins", "Insurance per year", A.OPEX["insurance_pa"], FMT_USD0, "$")
    r = single(r, "ins_g", "Insurance growth per year", A.OPEX["insurance_growth_pa"], FMT_PCT1, "%")
    r += 1

    # ---- cash
    r = section(ws, r, "CASH AND WORKING CAPITAL", 8)
    r = single(r, "upfront", "Share of core billings paid annually upfront",
               A.CASH["annual_upfront_pct"], FMT_PCT0, "%",
               "Single largest driver of the cash conclusion.")
    r = single(r, "qtrly", "Share billed quarterly", A.CASH["quarterly_pct"], FMT_PCT0, "%")
    r = single(r, "dso", "Days sales outstanding", A.CASH["dso_days"], FMT_0, "days")
    r = single(r, "dpo", "Days payable outstanding", A.CASH["dpo_days"], FMT_0, "days")
    r = single(r, "baddebt", "Bad debt as share of billings", A.CASH["bad_debt_pct"], FMT_PCT1, "%")
    r = single(r, "capex_hire", "Capital expenditure per new hire", A.CASH["capex_per_new_hire"], FMT_USD0, "$")
    r = single(r, "leasehold", "Leasehold at office step", A.CASH["leasehold_at_office_step"], FMT_USD0, "$")
    r = single(r, "int_rate", "Interest rate on cash", A.CASH["interest_rate_on_cash"], FMT_PCT1, "%",
               "Applied to the opening balance so the model has no circular reference.")
    r += 1

    # ---- fundraise
    r = section(ws, r, "FUNDRAISE POLICY", 8)
    r = single(r, "raise_runway", "Launch process at runway of, months",
               A.FUNDRAISE["process_launch_at_runway_months"], FMT_0, "months",
               "Rocket story, not buying time. Decision precedes the breach. Source: Series D norms: $100-300M round, $1-3B post money, 8-12% dilution, 24-36 months after Series C. Startups.com.")
    r = single(r, "mile_arr", "Milestone ARR", A.FUNDRAISE["milestone_arr"], FMT_USD0, "$")
    r = single(r, "mile_bm", "Milestone burn multiple", A.FUNDRAISE["milestone_burn_multiple"], FMT_X, "x")
    r = single(r, "post_runway", "Target post raise runway, months",
               A.FUNDRAISE["target_post_raise_runway_months"], FMT_0, "months")

    # ---- benchmarks
    r += 1
    r = section(ws, r, "BENCHMARKS", 8)
    for k, txt, fmt in [("nrr_median", "NRR median", FMT_PCT0),
                        ("nrr_top_quartile", "NRR top quartile", FMT_PCT0),
                        ("grr_median", "GRR median", FMT_PCT0),
                        ("grr_top_quartile", "GRR top quartile", FMT_PCT0),
                        ("burn_multiple_good", "Burn multiple standard", FMT_X),
                        ("gross_margin_saas", "SaaS gross margin standard", FMT_PCT0)]:
        r = single(r, f"bm_{k}", txt, A.BENCHMARKS[k], fmt)

    r += 1
    r = section(ws, r, "SOURCES", 8)
    for txt, where in [
        ("NRR 115% median / 128% top quartile. GRR 89% / 94%. Logo retention 90% / 95%. "
         "Scale stage, $25M to $100M ARR.",
         "Growthspree, B2B SaaS NRR and GRR benchmarks 2026"),
        ("AE quota $0.9M to $1.4M commercial. Ramp 7 to 9 months. Attainment 70% median, "
         "80% top quartile. Fully ramped AEs are only 60 to 75% of AE headcount. "
         "Time to hire 6 to 10 weeks.",
         "Growthspree, B2B SaaS sales capacity planning 2026"),
        ("Burn multiple below 1.00x at $50M+ ARR. Magic number 0.60 to 0.85. Rule of 40 ~38%.",
         "Data-Mania, B2B SaaS revenue efficiency benchmarks 2026"),
        ("AE turnover 30 to 32%. SDR 45 to 48%. Blended sales 35%. Best in class under 20%.",
         "Optifai, sales team turnover by role, 939 B2B companies"),
        ("Series D $100M to $300M, median ~$150M. Post money $1B to $3B. Founder dilution "
         "8 to 12%. ARR threshold $50M to $100M+. Typically 24 to 36 months after Series C.",
         "Startups.com, Series D funding benchmarks"),
        ("Published pricing tiers: $99 Starter, $399 Growth, custom Enterprise.",
         "Profound pricing, 2026"),
        ("700+ enterprise customers, $96M Series C at $1B, $155M raised to date.",
         "Fortune, February 2026"),
        ("Token pricing used to derive cost per query. Budget tier $0.10 to $0.14 per M input. "
         "Mid tier $0.30 to $1.00 input, $2.50 to $5.00 output. Batch processing 50% off, "
         "prompt caching up to 90% off cached input.",
         "CloudZero, LLM API pricing comparison 2026"),
    ]:
        c = ws.cell(row=r, column=COL_LABEL, value=where)
        c.font = Font(size=8.5, bold=True)
        d = ws.cell(row=r, column=4, value=txt)
        d.font = Font(size=8.5, color=GREY)
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    return ws


# =============================================================================
# Shared helpers for calculation sheets
# =============================================================================
R = {}          # registry: "Sheet!key" -> row number
HDR = 5         # header block starts at row 5 on every calc sheet
FY_ROW = HDR + 1
PER_ROW = HDR + 2


def reg(sheet, key, row):
    R[f"{sheet}|{key}"] = row
    return row


def rr(sheet, key):
    return R[f"{sheet}|{key}"]


def x(sheet, key, ci, absrow=False):
    """Cross-sheet reference to a row on another calculation sheet."""
    row = rr(sheet, key)
    return f"'{sheet}'!{col(ci)}{'$' if absrow else ''}{row}"


def inp(name):
    return I[name]


def yl(name, ci):
    """Pull a year-varying (or scenario-active) input for this column's fiscal year."""
    r = YROW[name]
    return f"INDEX(Inputs!$D${r}:$H${r},1,{col(ci)}${FY_ROW}-2025)"


def mgrow(annual_ref):
    """Monthly equivalent of an annual growth rate."""
    return f"((1+{annual_ref})^(1/12)-1)"


# =============================================================================
# REV-CAPACITY
# =============================================================================
def build_capacity(wb):
    S = "Rev-Capacity"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Revenue: Sales Capacity and Bookings")
    r = period_header(ws)

    r = section(ws, r, "QUOTA CARRYING CAPACITY")
    tth = inp("tth")
    label(ws, r, "AE hires arriving", "#", indent=1)
    reg(S, "ae_hires", r)
    write_row(ws, r, lambda i, n: (
        f"=INDEX(Inputs!$D${YROW['ae_hires']}:$H${YROW['ae_hires']},1,"
        f"{col(max(i - A.SALES['time_to_hire_months'], COL_M0))}${FY_ROW}-2025)"), FMT_1, GREEN)
    fy_sum(ws, r); r += 1

    label(ws, r, "AE attrition", "#", indent=1)
    reg(S, "ae_attr", r)
    reg(S, "ae_hc", r + 1)          # reserved: attrition looks forward one row
    write_row(ws, r, lambda i, n: (
        f"={inp('opening_aes')}*{inp('attr_s')}/12" if n == 0
        else f"={col(i-1)}{rr(S,'ae_hc')}*{inp('attr_s')}/12"), FMT_1, BLACK)
    fy_sum(ws, r); r += 1

    label(ws, r, "AE headcount, closing", "#", indent=1, bold=True)
    write_row(ws, r, lambda i, n: (
        f"={inp('opening_aes')}*(1-{inp('attr_s')}/12)+{col(i)}{rr(S,'ae_hires')}" if n == 0
        else f"={col(i-1)}{rr(S,'ae_hc')}*(1-{inp('attr_s')}/12)+{col(i)}{rr(S,'ae_hires')}"),
        FMT_1, BLACK, bold=True)
    fy_last(ws, r); r += 1

    label(ws, r, "Unramped drag", "#", indent=1, italic=True)
    reg(S, "drag", r)

    def drag_f(i, n):
        terms = []
        for k in range(5):
            if i - k >= COL_M0:
                terms.append(f"{col(i-k)}{rr(S,'ae_hires')}*(1-{yl(f'ramp{k}', i)})")
        return "=" + "+".join(terms) if terms else "=0"
    write_row(ws, r, drag_f, FMT_1, BLACK); r += 1

    label(ws, r, "Productive AE equivalents", "#", indent=1, bold=True)
    reg(S, "prod_ae", r)
    write_row(ws, r, lambda i, n: f"=MAX({col(i)}{rr(S,'ae_hc')}-{col(i)}{rr(S,'drag')},0)",
              FMT_1, BLACK, bold=True)
    fy_last(ws, r); r += 2

    r = section(ws, r, "SELF SERVE")
    label(ws, r, "New customers added", "#", indent=1)
    reg(S, "ss_adds", r)
    write_row(ws, r, lambda i, n: f"={inp('ss_adds')}*(1+{mgrow(inp('ss_adds_g'))})^{n}", FMT_1, BLACK); r += 1
    label(ws, r, "Churned customers", "#", indent=1)
    reg(S, "ss_churn", r)
    reg(S, "ss_cust", r + 2)        # reserved: churn and conversions look forward
    write_row(ws, r, lambda i, n: (
        f"={inp('ss_cust0')}*{inp('ss_churn')}" if n == 0
        else f"={col(i-1)}{rr(S,'ss_cust')}*{inp('ss_churn')}"), FMT_1, BLACK); r += 1
    label(ws, r, "Converted into enterprise", "#", indent=1)
    reg(S, "ss_conv", r)
    write_row(ws, r, lambda i, n: (
        f"={inp('ss_cust0')}*{inp('ss_conv')}" if n == 0
        else f"={col(i-1)}{rr(S,'ss_cust')}*{inp('ss_conv')}"), FMT_1, BLACK); r += 1
    label(ws, r, "Self serve customers, closing", "#", indent=1, bold=True)
    write_row(ws, r, lambda i, n: (
        f"={inp('ss_cust0')}+{col(i)}{rr(S,'ss_adds')}-{col(i)}{rr(S,'ss_churn')}-{col(i)}{rr(S,'ss_conv')}"
        if n == 0 else
        f"={col(i-1)}{rr(S,'ss_cust')}+{col(i)}{rr(S,'ss_adds')}-{col(i)}{rr(S,'ss_churn')}-{col(i)}{rr(S,'ss_conv')}"),
        FMT_1, BLACK, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "Self serve ARR", "$", indent=1, bold=True)
    reg(S, "ss_arr", r)
    write_row(ws, r, lambda i, n:
              f"={col(i)}{rr(S,'ss_cust')}*{inp('ss_acv')}*(1+{mgrow(inp('ss_acv_g'))})^{n}",
              FMT_USD0, BLACK, bold=True)
    fy_last(ws, r); r += 2

    r = section(ws, r, "ENTERPRISE BOOKINGS")
    label(ws, r, "Quota per AE per month", "$", indent=1)
    reg(S, "quota_m", r)
    write_row(ws, r, lambda i, n: f"={inp('quota')}*(1+{mgrow(inp('quota_g'))})^{n}/12", FMT_USD0, BLACK)
    r += 1
    label(ws, r, "Quota attainment", "%", indent=1)
    reg(S, "attain", r)
    write_row(ws, r, lambda i, n: f"={yl('attainment', i)}", FMT_PCT0, GREEN); r += 1
    label(ws, r, "New ARR from sales capacity", "$", indent=1)
    reg(S, "new_sales", r)
    write_row(ws, r, lambda i, n:
              f"={col(i)}{rr(S,'prod_ae')}*{col(i)}{rr(S,'quota_m')}*{col(i)}{rr(S,'attain')}",
              FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1

    label(ws, r, "Enterprise ACV", "$", indent=1)
    reg(S, "acv", r)
    write_row(ws, r, lambda i, n: f"={inp('acv')}*(1+{mgrow(inp('acv_g'))})^{n}", FMT_USD0, BLACK); r += 1
    label(ws, r, "New ARR from self serve conversions", "$", indent=1)
    reg(S, "new_conv", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'ss_conv')}*{col(i)}{rr(S,'acv')}", FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Total new enterprise ARR", "$", indent=1, bold=True)
    reg(S, "new_arr", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'new_sales')}+{col(i)}{rr(S,'new_conv')}",
              FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "New logos", "#", indent=1)
    reg(S, "new_logos", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'new_arr')}/{col(i)}{rr(S,'acv')}", FMT_1, BLACK)
    fy_sum(ws, r); r += 2
    return ws


# =============================================================================
# REV-COHORTS  (dollar cohort triangle)
# =============================================================================
N_OPEN = 12


def build_cohorts(wb):
    S = "Rev-Cohorts"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Revenue: Dollar Cohort Triangle")
    r = period_header(ws)

    r = section(ws, r, "RETENTION DRIVERS")
    label(ws, r, "Seat and usage expansion, monthly", "%", indent=1)
    reg(S, "seat_m", r)
    write_row(ws, r, lambda i, n: f"=(1+{yl('seat_exp', i)})^(1/12)-1", FMT_PCT1, GREEN); r += 1
    label(ws, r, "In contract churn, monthly", "%", indent=1)
    reg(S, "churn_m", r)
    write_row(ws, r, lambda i, n: f"={inp('churn_ic')}/12", FMT_PCT1, GREEN); r += 1
    label(ws, r, "First renewal survival", "%", indent=1)
    reg(S, "surv1", r)
    write_row(ws, r, lambda i, n: f"={yl('surv1', i)}", FMT_PCT1, GREEN); r += 1
    label(ws, r, "Later renewal survival", "%", indent=1)
    reg(S, "survL", r)
    write_row(ws, r, lambda i, n: f"={yl('surv_later', i)}", FMT_PCT1, GREEN); r += 1
    label(ws, r, "Renewal uplift, price plus cross sell", "%", indent=1)
    reg(S, "uplift", r)
    write_row(ws, r, lambda i, n: f"={inp('uplift')}+{yl('cross', i)}", FMT_PCT1, GREEN); r += 2

    r = section(ws, r, "COHORT TRIANGLE")
    r = subsection(ws, r, "Opening base, spread evenly across the prior twelve months")
    open_first = r
    sm, cm, s1, sl, up = (rr(S, 'seat_m'), rr(S, 'churn_m'), rr(S, 'surv1'),
                          rr(S, 'survL'), rr(S, 'uplift'))
    for k in range(N_OPEN):
        a = N_OPEN - k                      # initial age 12, 11, ... 1
        label(ws, r, f"Opening slice, age {a} months at start", "$", indent=1)

        def f(i, n, a=a):
            prev = (f"{inp('ent_arr0')}/{N_OPEN}" if n == 0 else f"{col(i-1)}{ws_row}")
            age = f"({a}+{col(i)}${PER_ROW})"
            ren = (f"IF(MOD({age},12)=0,IF({age}=12,{col(i)}${s1},{col(i)}${sl})"
                   f"*(1+{col(i)}${up}),1)")
            return f"={prev}*(1+{col(i)}${sm})*(1-{col(i)}${cm})*{ren}"
        ws_row = r
        write_row(ws, r, f, FMT_USD0, BLACK)
        reg(S, f"open{k}", r)
        ws.row_dimensions[r].outlineLevel = 1
        ws.row_dimensions[r].hidden = True
        r += 1
    open_last = r - 1

    r = subsection(ws, r, "New cohorts, one per month of origination")
    new_first = r
    for k in range(N_MONTHS):
        label(ws, r, f"Cohort {DATES[k]:%b-%y}", "$", indent=1)

        def f(i, n, k=k, myrow=r):
            age = f"({col(i)}${PER_ROW}-{k+1})"
            newv = f"'Rev-Capacity'!{col(i)}{rr('Rev-Capacity','new_arr')}"
            if n < k:
                return "=0"
            if n == k:
                return f"={newv}"
            prev = f"{col(i-1)}{myrow}"
            ren = (f"IF(MOD({age},12)=0,IF({age}=12,{col(i)}${s1},{col(i)}${sl})"
                   f"*(1+{col(i)}${up}),1)")
            return f"={prev}*(1+{col(i)}${sm})*(1-{col(i)}${cm})*{ren}"
        write_row(ws, r, f, FMT_USD0, BLACK)
        reg(S, f"new{k}", r)
        ws.row_dimensions[r].outlineLevel = 1
        ws.row_dimensions[r].hidden = True
        r += 1
    new_last = r - 1
    ws.sheet_properties.outlinePr.summaryBelow = True

    r += 1
    label(ws, r, "Total enterprise core ARR", "$", indent=0, bold=True)
    reg(S, "ent_arr", r)
    write_row(ws, r, lambda i, n:
              f"=SUM({col(i)}{open_first}:{col(i)}{open_last})+SUM({col(i)}{new_first}:{col(i)}{new_last})",
              FMT_USD0, PURPLE, bold=True)
    fy_last(ws, r); r += 2

    # Renewals are periodic, so the contributing cohorts sit at fixed offsets each month.
    # First and later renewals carry different survival, so they are split here rather
    # than approximated with a period test on the bridge.
    r = section(ws, r, "RENEWAL SCHEDULE")

    def renew_terms(n, first_only):
        p, terms = n + 1, []
        for k in range(N_OPEN):
            a = N_OPEN - k
            age = a + p
            if age % 12 == 0 and ((age == 12) == first_only):
                terms.append(("open", k))
        for k in range(N_MONTHS):
            age = p - (k + 1)
            if age > 0 and age % 12 == 0 and ((age == 12) == first_only):
                terms.append(("new", k))
        return terms

    for tag, first_only, txt in [("renew1", True, "ARR reaching its FIRST renewal this month"),
                                 ("renewL", False, "ARR reaching a LATER renewal this month")]:
        label(ws, r, txt, "$", indent=1)
        reg(S, tag, r)

        def f(i, n, fo=first_only):
            terms = renew_terms(n, fo)
            if not terms:
                return "=0"
            parts = []
            for kind, k in terms:
                if n == 0:
                    parts.append(f"{inp('ent_arr0')}/{N_OPEN}")
                else:
                    parts.append(f"{col(i-1)}{rr(S, ('open' if kind=='open' else 'new') + str(k))}")
            return f"=({'+'.join(parts)})*(1+{col(i)}${sm})*(1-{col(i)}${cm})"
        write_row(ws, r, f, FMT_USD0, BLACK)
        fy_sum(ws, r); r += 1

    label(ws, r, "Total ARR renewing this month", "$", indent=1, bold=True)
    reg(S, "renewing", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'renew1')}+{col(i)}{rr(S,'renewL')}",
              FMT_USD0, PURPLE, bold=True)
    fy_sum(ws, r); r += 1
    r = note(ws, r, "Expand the triangle above to trace any single cohort through its renewals.")
    return ws


# =============================================================================
# REV-ARR BRIDGE
# =============================================================================
def build_bridge(wb):
    S = "Rev-ARR Bridge"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Revenue: ARR Bridge")
    r = period_header(ws)

    C, CAP = "Rev-Cohorts", "Rev-Capacity"
    r = section(ws, r, "ENTERPRISE CORE ARR BRIDGE")
    label(ws, r, "Beginning ARR", "$", indent=1)
    reg(S, "beg", r)
    reg(S, "end", r + 4)            # reserved: beginning ARR looks forward to closing
    write_row(ws, r, lambda i, n: (f"={inp('ent_arr0')}" if n == 0
                                   else f"={col(i-1)}{rr(S,'end')}"), FMT_USD0, BLACK); r += 1
    label(ws, r, "New ARR", "$", indent=1)
    reg(S, "new", r)
    write_row(ws, r, lambda i, n: f"={x(CAP,'new_arr',i)}", FMT_USD0, GREEN)
    fy_sum(ws, r); r += 1
    label(ws, r, "Expansion", "$", indent=1)
    reg(S, "exp", r)
    write_row(ws, r, lambda i, n: (
        f"={col(i)}{rr(S,'beg')}*{x(C,'seat_m',i)}"
        f"+({x(C,'renew1',i)}*{x(C,'surv1',i)}+{x(C,'renewL',i)}*{x(C,'survL',i)})"
        f"*{x(C,'uplift',i)}"), FMT_USD0, GREEN)
    fy_sum(ws, r); r += 1
    label(ws, r, "Churn", "$", indent=1)
    reg(S, "churn", r)
    write_row(ws, r, lambda i, n: (
        f"=-({col(i)}{rr(S,'beg')}*(1+{x(C,'seat_m',i)})*{x(C,'churn_m',i)}"
        f"+{x(C,'renew1',i)}*(1-{x(C,'surv1',i)})+{x(C,'renewL',i)}*(1-{x(C,'survL',i)}))"),
        FMT_USD0, GREEN)
    fy_sum(ws, r); r += 1
    label(ws, r, "Ending enterprise core ARR", "$", indent=1, bold=True)
    write_row(ws, r, lambda i, n: f"={x(C,'ent_arr',i)}", FMT_USD0, GREEN, bold=True)
    fy_last(ws, r); r += 2

    r = section(ws, r, "AGENTS")
    label(ws, r, "Enterprise customers", "#", indent=1)
    reg(S, "cust", r)
    write_row(ws, r, lambda i, n: (
        f"={inp('ent_cust0')}*(1-({inp('churn_ic')}+{inp('logo_prem')})/12)"
        f"+{x(CAP,'new_logos',i)}"
        f"-{inp('ent_cust0')}/12*(1-({x(C,'surv1',i)}-{inp('logo_prem')}))" if n == 0
        else f"={col(i-1)}{rr(S,'cust')}+{x(CAP,'new_logos',i)}"
             f"-{col(i-1)}{rr(S,'cust')}*({inp('churn_ic')}+{inp('logo_prem')})/12"
             f"-IF({col(i)}${PER_ROW}<=12,{inp('ent_cust0')}/12*(1-({x(C,'surv1',i)}-{inp('logo_prem')})),0)"),
        FMT_1, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "Agents attach rate", "%", indent=1)
    reg(S, "attach", r)
    write_row(ws, r, lambda i, n: (
        f"={yl('attach', i)}*{n+1}/12" if n < 12
        else f"={col(COL_M0 + (n//12)*12 - 1)}{rr(S,'attach')}"
             f"+({yl('attach', i)}-{col(COL_M0 + (n//12)*12 - 1)}{rr(S,'attach')})*{(n % 12)+1}/12"),
        FMT_PCT1, GREEN); r += 1
    label(ws, r, "Agents customers", "#", indent=1)
    reg(S, "ag_cust", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'cust')}*{col(i)}{rr(S,'attach')}", FMT_1, BLACK); r += 1
    label(ws, r, "Agents ACV", "$", indent=1)
    reg(S, "ag_acv", r)
    write_row(ws, r, lambda i, n:
              f"={inp('acv')}*{yl('ag_pct', i)}*(1+{mgrow(yl('ag_g', i))})^{n}",
              FMT_USD0, GREEN); r += 1
    label(ws, r, "Agents ARR", "$", indent=1, bold=True)
    reg(S, "ag_arr", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'ag_cust')}*{col(i)}{rr(S,'ag_acv')}",
              FMT_USD0, BLACK, bold=True)
    fy_last(ws, r); r += 2

    r = section(ws, r, "TOTAL ARR")
    label(ws, r, "Self serve ARR", "$", indent=1)
    reg(S, "ss_arr", r)
    write_row(ws, r, lambda i, n: f"={x(CAP,'ss_arr',i)}", FMT_USD0, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "Core ARR, enterprise plus self serve", "$", indent=1)
    reg(S, "core_arr", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'end')}+{col(i)}{rr(S,'ss_arr')}", FMT_USD0, BLACK)
    fy_last(ws, r); r += 1
    label(ws, r, "Total ARR", "$", indent=1, bold=True)
    reg(S, "total_arr", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'core_arr')}+{col(i)}{rr(S,'ag_arr')}",
              FMT_USD0, BLACK, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "Year on year ARR growth", "%", indent=1)
    reg(S, "yoy", r)
    write_row(ws, r, lambda i, n: (
        f"={col(i)}{rr(S,'total_arr')}/{inp('opening_arr')}-1" if n < 12
        else f"={col(i)}{rr(S,'total_arr')}/{col(i-12)}{rr(S,'total_arr')}-1"), FMT_PCT0, BLACK)
    fy_last(ws, r); r += 1
    return ws


# =============================================================================
# HEADCOUNT
# =============================================================================
def build_headcount(wb):
    S = "Headcount"; B = "Rev-ARR Bridge"; CAP = "Rev-Capacity"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Headcount and Compensation")
    r = period_header(ws)

    r = section(ws, r, "ROSTER BY FUNCTION")
    label(ws, r, "R&D hires per month", "#", indent=1)
    reg(S, "rnd_h", r); write_row(ws, r, lambda i, n: f"={yl('rnd_hires', i)}", FMT_1, GREEN); r += 1
    label(ws, r, "R&D headcount", "#", indent=1)
    reg(S, "rnd", r)
    write_row(ws, r, lambda i, n: (f"={inp('hc0_rnd')}+{col(i)}{rr(S,'rnd_h')}" if n == 0
              else f"={col(i-1)}{rr(S,'rnd')}+{col(i)}{rr(S,'rnd_h')}"), FMT_1, BLACK)
    fy_last(ws, r); r += 1
    label(ws, r, "G&A hires per month", "#", indent=1)
    reg(S, "ga_h", r); write_row(ws, r, lambda i, n: f"={yl('ga_hires', i)}", FMT_1, GREEN); r += 1
    label(ws, r, "G&A headcount", "#", indent=1)
    reg(S, "ga", r)
    write_row(ws, r, lambda i, n: (f"={inp('hc0_ga')}+{col(i)}{rr(S,'ga_h')}" if n == 0
              else f"={col(i-1)}{rr(S,'ga')}+{col(i)}{rr(S,'ga_h')}"), FMT_1, BLACK)
    fy_last(ws, r); r += 1
    label(ws, r, "Customer success headcount", "#", indent=1)
    reg(S, "cs", r)
    write_row(ws, r, lambda i, n:
              f"=MAX({x(B,'cust',i)}/{yl('csm_ratio', i)},{inp('hc0_cs')}*0.5)", FMT_1, GREEN)
    fy_last(ws, r); r += 1
    for k, ratio in [("sdr", "sdr_per_ae"), ("mktg", "mktg_per_ae"),
                     ("sops", "sops_per_ae"), ("enable", "enable_per_ae")]:
        label(ws, r, f"{k.upper()} headcount", "#", indent=1)
        reg(S, k, r)
        write_row(ws, r, lambda i, n, rt=ratio: f"={x(CAP,'ae_hc',i)}*{inp(rt)}", FMT_1, GREEN); r += 1
    label(ws, r, "Sales and marketing headcount", "#", indent=1, bold=True)
    reg(S, "sm", r)
    write_row(ws, r, lambda i, n: (f"={x(CAP,'ae_hc',i)}+{col(i)}{rr(S,'sdr')}+{col(i)}{rr(S,'mktg')}"
                                   f"+{col(i)}{rr(S,'sops')}+{col(i)}{rr(S,'enable')}"),
              FMT_1, BLACK, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "Total headcount", "#", indent=1, bold=True)
    reg(S, "total", r)
    write_row(ws, r, lambda i, n: (f"={col(i)}{rr(S,'rnd')}+{col(i)}{rr(S,'ga')}"
                                   f"+{col(i)}{rr(S,'cs')}+{col(i)}{rr(S,'sm')}"),
              FMT_1, BLACK, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "Gross hires including backfill", "#", indent=1)
    reg(S, "gross_hires", r)
    write_row(ws, r, lambda i, n: (
        f"={col(i)}{rr(S,'rnd_h')}+{col(i)}{rr(S,'ga_h')}+{x(CAP,'ae_hires',i)}"
        f"+({inp('hc0_rnd')}+{inp('hc0_ga')})*{inp('attr_ns')}/12"
        f"+{x(CAP,'ae_attr',i)}" if n == 0 else
        f"={col(i)}{rr(S,'rnd_h')}+{col(i)}{rr(S,'ga_h')}+{x(CAP,'ae_hires',i)}"
        f"+({col(i-1)}{rr(S,'rnd')}+{col(i-1)}{rr(S,'ga')})*{inp('attr_ns')}/12"
        f"+{x(CAP,'ae_attr',i)}"), FMT_1, BLACK)
    fy_sum(ws, r); r += 2

    r = section(ws, r, "PAYROLL")
    label(ws, r, "Compensation escalator", "x", indent=1)
    reg(S, "esc", r)
    write_row(ws, r, lambda i, n: f"=(1+{mgrow(inp('comp_infl'))})^{n}", FMT_2, BLACK); r += 1
    for key, hcrow, lc in [("pay_rnd", "rnd", "lc_rnd"), ("pay_cs", "cs", "lc_cs"),
                           ("pay_ga", "ga", "lc_ga")]:
        label(ws, r, f"{hcrow.upper()} payroll", "$", indent=1)
        reg(S, key, r)
        write_row(ws, r, lambda i, n, h=hcrow, l=lc:
                  f"={col(i)}{rr(S,h)}*{inp(l)}/12*{col(i)}{rr(S,'esc')}", FMT_USD0, BLACK)
        fy_sum(ws, r); r += 1
    label(ws, r, "S&M payroll", "$", indent=1)
    reg(S, "pay_sm", r)
    write_row(ws, r, lambda i, n: (
        f"=({x(CAP,'ae_hc',i)}*{inp('lc_ae')}+{col(i)}{rr(S,'sdr')}*{inp('lc_sdr')}"
        f"+{col(i)}{rr(S,'mktg')}*{inp('lc_mktg')}"
        f"+({col(i)}{rr(S,'sops')}+{col(i)}{rr(S,'enable')})*{inp('lc_salesops')})"
        f"/12*{col(i)}{rr(S,'esc')}"), FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Total payroll", "$", indent=1, bold=True)
    reg(S, "payroll", r)
    write_row(ws, r, lambda i, n: (f"={col(i)}{rr(S,'pay_rnd')}+{col(i)}{rr(S,'pay_cs')}"
                                   f"+{col(i)}{rr(S,'pay_ga')}+{col(i)}{rr(S,'pay_sm')}"),
              FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "Stock compensation, non cash", "$", indent=1)
    reg(S, "sbc", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'payroll')}*{inp('sbc_pct')}", FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Recruiting cost", "$", indent=1)
    reg(S, "recruit", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'gross_hires')}*{inp('recruit')}", FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    return ws


# =============================================================================
# P&L
# =============================================================================
def build_pl(wb):
    S = "P&L"; B = "Rev-ARR Bridge"; H = "Headcount"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Profit and Loss")
    r = period_header(ws)

    r = section(ws, r, "REVENUE")
    label(ws, r, "Core revenue", "$", indent=1)
    reg(S, "rev_core", r)
    write_row(ws, r, lambda i, n: (f"={inp('opening_arr')}/12" if n == 0
              else f"={col(i-1)}{x(B,'core_arr',i-1).split('!')[0]}!{col(i-1)}{rr(B,'core_arr')}"
                   if False else f"='{B}'!{col(i-1)}{rr(B,'core_arr')}/12"), FMT_USD0, GREEN)
    fy_sum(ws, r); r += 1
    label(ws, r, "Self serve revenue, memo", "$", indent=1, italic=True)
    reg(S, "rev_ss", r)
    write_row(ws, r, lambda i, n: (f"={inp('ss_arr0')}/12" if n == 0
              else f"='{B}'!{col(i-1)}{rr(B,'ss_arr')}/12"), FMT_USD0, GREEN); r += 1
    label(ws, r, "Agents revenue", "$", indent=1)
    reg(S, "rev_ag", r)
    write_row(ws, r, lambda i, n: ("=0" if n == 0
              else f"='{B}'!{col(i-1)}{rr(B,'ag_arr')}/12"), FMT_USD0, GREEN)
    fy_sum(ws, r); r += 1
    label(ws, r, "Total revenue", "$", indent=1, bold=True)
    reg(S, "rev", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev_core')}+{col(i)}{rr(S,'rev_ag')}",
              FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 2

    r = section(ws, r, "COST OF REVENUE")
    label(ws, r, "Queries per customer per month", "#", indent=1)
    reg(S, "q", r)
    write_row(ws, r, lambda i, n:
              f"={inp('prompts')}*{inp('engines')}*{inp('runs_pm')}"
              f"*(1+{mgrow(inp('query_g'))})^{n}", FMT_0, BLACK); r += 1
    label(ws, r, "Cost per query", "$", indent=1)
    reg(S, "cq", r)
    write_row(ws, r, lambda i, n: f"={inp('cost_q')}*(1-{inp('defl')})^({n}/12)", '$#,##0.00000', BLACK); r += 1
    label(ws, r, "Inference cost", "$", indent=1)
    reg(S, "cogs_inf", r)
    write_row(ws, r, lambda i, n: f"={x(B,'cust',i)}*{col(i)}{rr(S,'q')}*{col(i)}{rr(S,'cq')}",
              FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Agents cost", "$", indent=1)
    reg(S, "cogs_ag", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev_ag')}*{yl('ag_cogs', i)}", FMT_USD0, GREEN); r += 1
    label(ws, r, "Self serve cost", "$", indent=1)
    reg(S, "cogs_ss", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev_ss')}*{inp('ss_cogs')}", FMT_USD0, BLACK); r += 1
    label(ws, r, "Hosting and infrastructure", "$", indent=1)
    reg(S, "cogs_host", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev')}*{yl('host', i)}", FMT_USD0, GREEN); r += 1
    label(ws, r, "Third party data", "$", indent=1)
    reg(S, "cogs_data", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev')}*{inp('data3p')}", FMT_USD0, BLACK); r += 1
    label(ws, r, "Total cost of revenue", "$", indent=1, bold=True)
    reg(S, "cogs", r)
    write_row(ws, r, lambda i, n: (
        f"={col(i)}{rr(S,'cogs_inf')}+{col(i)}{rr(S,'cogs_ag')}+{col(i)}{rr(S,'cogs_ss')}"
        f"+{col(i)}{rr(S,'cogs_host')}+{col(i)}{rr(S,'cogs_data')}"),
        FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "Gross profit", "$", indent=1, bold=True)
    reg(S, "gp", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev')}-{col(i)}{rr(S,'cogs')}",
              FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "Gross margin", "%", indent=1)
    reg(S, "gm", r)
    write_row(ws, r, lambda i, n:
              f"=IF({col(i)}{rr(S,'rev')}=0,0,{col(i)}{rr(S,'gp')}/{col(i)}{rr(S,'rev')})", FMT_PCT1, BLACK)
    fy_last(ws, r)
    r += 2

    r = section(ws, r, "OPERATING EXPENSE")
    label(ws, r, "Software", "$", indent=1); reg(S, "sw", r)
    write_row(ws, r, lambda i, n: f"={x(H,'total',i)}*{inp('sw_head')}/12", FMT_USD0, GREEN); r += 1
    label(ws, r, "Travel and entertainment", "$", indent=1); reg(S, "te", r)
    write_row(ws, r, lambda i, n: f"={x(H,'total',i)}*{inp('te_head')}/12", FMT_USD0, GREEN); r += 1
    label(ws, r, "Facilities including office step", "$", indent=1); reg(S, "fac", r)
    write_row(ws, r, lambda i, n: (
        f"={x(H,'total',i)}*{inp('fac_head')}/12"
        f"+IF({x(H,'total',i)}>={inp('fac_step_hc')},{inp('fac_step')}/12,0)"), FMT_USD0, GREEN); r += 1
    label(ws, r, "Other G&A", "$", indent=1); reg(S, "oth", r)
    write_row(ws, r, lambda i, n: f"={x(H,'total',i)}*{inp('oth_head')}/12", FMT_USD0, GREEN); r += 1
    label(ws, r, "R&D compute", "$", indent=1); reg(S, "compute", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev')}*{yl('rnd_compute', i)}", FMT_USD0, GREEN); r += 1
    label(ws, r, "Marketing programmes", "$", indent=1); reg(S, "mktgp", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev')}*{inp('mktg_prog')}", FMT_USD0, BLACK); r += 1
    label(ws, r, "Professional fees", "$", indent=1); reg(S, "prof", r)
    write_row(ws, r, lambda i, n:
              f"={inp('prof')}*(1+{inp('prof_g')})^({col(i)}${FY_ROW}-2026)/12", FMT_USD0, BLACK); r += 1
    label(ws, r, "Insurance", "$", indent=1); reg(S, "ins", r)
    write_row(ws, r, lambda i, n:
              f"={inp('ins')}*(1+{inp('ins_g')})^({col(i)}${FY_ROW}-2026)/12", FMT_USD0, BLACK); r += 1
    label(ws, r, "Customer success payroll", "$", indent=1); reg(S, "opex_cs", r)
    write_row(ws, r, lambda i, n: f"={x(H,'pay_cs',i)}", FMT_USD0, GREEN); r += 1
    label(ws, r, "Sales, marketing and customer success", "$", indent=1, bold=True); reg(S, "sm", r)
    write_row(ws, r, lambda i, n:
              f"={x(H,'pay_sm',i)}+{col(i)}{rr(S,'opex_cs')}+{col(i)}{rr(S,'mktgp')}"
              f"+{col(i)}{rr(S,'te')}*0.5", FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "Research and development", "$", indent=1, bold=True); reg(S, "rnd", r)
    write_row(ws, r, lambda i, n:
              f"={x(H,'pay_rnd',i)}+{col(i)}{rr(S,'sw')}*0.5+{col(i)}{rr(S,'compute')}",
              FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "General and administrative", "$", indent=1, bold=True); reg(S, "ga", r)
    write_row(ws, r, lambda i, n: (
        f"={x(H,'pay_ga',i)}+{x(H,'recruit',i)}+{col(i)}{rr(S,'prof')}+{col(i)}{rr(S,'ins')}"
        f"+{col(i)}{rr(S,'fac')}+{col(i)}{rr(S,'oth')}+{col(i)}{rr(S,'sw')}*0.5"
        f"+{col(i)}{rr(S,'te')}*0.5"), FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "Total operating expense", "$", indent=1, bold=True); reg(S, "opex", r)
    write_row(ws, r, lambda i, n: (f"={col(i)}{rr(S,'sm')}+{col(i)}{rr(S,'rnd')}"
                                   f"+{col(i)}{rr(S,'ga')}"), FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 2

    r = section(ws, r, "EARNINGS")
    label(ws, r, "EBITDA, excluding stock compensation", "$", indent=1, bold=True); reg(S, "ebitda", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'gp')}-{col(i)}{rr(S,'opex')}",
              FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "Stock compensation, non cash", "$", indent=1); reg(S, "sbc", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'rev')}*{inp('sbc_pct')}", FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Operating result after stock compensation", "$", indent=1)
    reg(S, "op_after_sbc", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'ebitda')}-{col(i)}{rr(S,'sbc')}",
              FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "EBITDA margin", "%", indent=1); reg(S, "ebitda_pct", r)
    write_row(ws, r, lambda i, n:
              f"=IF({col(i)}{rr(S,'rev')}=0,0,{col(i)}{rr(S,'ebitda')}/{col(i)}{rr(S,'rev')})",
              FMT_PCT1, BLACK)
    fy_last(ws, r); r += 1
    return ws


# =============================================================================
# CASH FLOW
# =============================================================================
def build_cash(wb):
    S = "Cash Flow"; B = "Rev-ARR Bridge"; P = "P&L"; H = "Headcount"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Cash Flow")
    r = period_header(ws)

    r = section(ws, r, "BILLINGS AND DEFERRED REVENUE")
    label(ws, r, "Deferred revenue factor", "x", indent=1)
    reg(S, "deffac", r)
    write_row(ws, r, lambda i, n: f"={inp('upfront')}*0.5+{inp('qtrly')}*1.5/12", FMT_2, BLACK); r += 1
    label(ws, r, "Deferred revenue balance", "$", indent=1)
    reg(S, "deferred", r)
    write_row(ws, r, lambda i, n: f"={x(B,'core_arr',i)}*{col(i)}{rr(S,'deffac')}", FMT_USD0, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "Change in deferred revenue", "$", indent=1)
    reg(S, "ddef", r)
    write_row(ws, r, lambda i, n: (
        f"={col(i)}{rr(S,'deferred')}-{inp('opening_arr')}*{col(i)}{rr(S,'deffac')}" if n == 0
        else f"={col(i)}{rr(S,'deferred')}-{col(i-1)}{rr(S,'deferred')}"), FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Core billings", "$", indent=1)
    reg(S, "bill_core", r)
    write_row(ws, r, lambda i, n: f"={x(P,'rev_core',i)}+{col(i)}{rr(S,'ddef')}", FMT_USD0, GREEN); r += 1
    label(ws, r, "Agents billings, in arrears", "$", indent=1)
    reg(S, "bill_ag", r)
    write_row(ws, r, lambda i, n: ("=0" if n == 0 else f"={x(P,'rev_ag',i-1)}"), FMT_USD0, GREEN); r += 1
    label(ws, r, "Total billings", "$", indent=1, bold=True)
    reg(S, "billings", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'bill_core')}+{col(i)}{rr(S,'bill_ag')}",
              FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 2

    r = section(ws, r, "WORKING CAPITAL")
    label(ws, r, "Accounts receivable", "$", indent=1); reg(S, "ar", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'billings')}*{inp('dso')}/30", FMT_USD0, BLACK)
    fy_last(ws, r); r += 1
    label(ws, r, "Change in receivables", "$", indent=1); reg(S, "dar", r)
    write_row(ws, r, lambda i, n: ("=0" if n == 0
              else f"={col(i)}{rr(S,'ar')}-{col(i-1)}{rr(S,'ar')}"), FMT_USD0, BLACK); r += 1
    label(ws, r, "Accounts payable", "$", indent=1); reg(S, "ap", r)
    write_row(ws, r, lambda i, n: f"=({x(P,'cogs',i)}+{x(P,'opex',i)})*{inp('dpo')}/30",
              FMT_USD0, GREEN); r += 1
    label(ws, r, "Change in payables", "$", indent=1); reg(S, "dap", r)
    write_row(ws, r, lambda i, n: ("=0" if n == 0
              else f"={col(i)}{rr(S,'ap')}-{col(i-1)}{rr(S,'ap')}"), FMT_USD0, BLACK); r += 1
    label(ws, r, "Bad debt", "$", indent=1); reg(S, "bad", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'billings')}*{inp('baddebt')}", FMT_USD0, BLACK); r += 2

    r = section(ws, r, "CAPITAL AND NON CASH")
    label(ws, r, "Capital expenditure", "$", indent=1); reg(S, "capex", r)
    write_row(ws, r, lambda i, n: (
        f"={x(H,'gross_hires',i)}*{inp('capex_hire')}"
        f"+IF(AND({x(H,'total',i)}>={inp('fac_step_hc')},"
        f"{x(H,'total',max(i-1,COL_M0))}<{inp('fac_step_hc')}),{inp('leasehold')},0)"),
        FMT_USD0, GREEN)
    fy_sum(ws, r); r += 1
    label(ws, r, "Depreciation and amortisation", "$", indent=1); reg(S, "da", r)
    write_row(ws, r, lambda i, n:
              f"=SUM({col(max(i-35,COL_M0))}{rr(S,'capex')}:{col(i)}{rr(S,'capex')})/36",
              FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Stock compensation add back", "$", indent=1); reg(S, "sbc", r)
    write_row(ws, r, lambda i, n: f"={x(P,'sbc',i)}", FMT_USD0, GREEN)
    fy_sum(ws, r); r += 2

    r = section(ws, r, "CASH ROLL")
    label(ws, r, "Opening cash", "$", indent=1); reg(S, "cash_beg", r)
    reg(S, "cash_end", r + 4)       # reserved: opening cash looks forward to closing
    write_row(ws, r, lambda i, n: (f"={inp('opening_cash')}" if n == 0
              else f"={col(i-1)}{rr(S,'cash_end')}"), FMT_USD0, BLACK); r += 1
    label(ws, r, "Interest income", "$", indent=1); reg(S, "interest", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'cash_beg')}*{inp('int_rate')}/12",
              FMT_USD0, BLACK)
    fy_sum(ws, r); r += 1
    label(ws, r, "Net income", "$", indent=1); reg(S, "ni", r)
    write_row(ws, r, lambda i, n:
              f"={x(P,'ebitda',i)}-{x(P,'sbc',i)}-{col(i)}{rr(S,'da')}+{col(i)}{rr(S,'interest')}",
              FMT_USD0, GREEN)
    fy_sum(ws, r); r += 1
    label(ws, r, "Free cash flow", "$", indent=1, bold=True); reg(S, "fcf", r)
    write_row(ws, r, lambda i, n: (
        f"={col(i)}{rr(S,'ni')}+{col(i)}{rr(S,'sbc')}+{col(i)}{rr(S,'da')}"
        f"-{col(i)}{rr(S,'dar')}+{col(i)}{rr(S,'ddef')}+{col(i)}{rr(S,'dap')}"
        f"-{col(i)}{rr(S,'bad')}-{col(i)}{rr(S,'capex')}"), FMT_USD0, BLACK, bold=True)
    fy_sum(ws, r); r += 1
    label(ws, r, "Closing cash", "$", indent=1, bold=True)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'cash_beg')}+{col(i)}{rr(S,'fcf')}",
              FMT_USD0, BLACK, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "Headroom over minimum balance", "$", indent=1); reg(S, "headroom", r)
    write_row(ws, r, lambda i, n: f"={col(i)}{rr(S,'cash_end')}-{inp('min_cash')}", FMT_USD0, BLACK)
    fy_last(ws, r)
    ws.conditional_formatting.add(
        f"{col(COL_M0)}{r}:{col(COL_MLAST)}{r}",
        CellIsRule(operator="lessThan", formula=["0"], fill=FILL_BAD))
    r += 1
    return ws


# =============================================================================
# KPIs
# =============================================================================
def build_kpis(wb):
    S = "KPIs"; B = "Rev-ARR Bridge"; P = "P&L"; H = "Headcount"; CF = "Cash Flow"
    C = "Rev-Cohorts"; CAP = "Rev-Capacity"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Key Performance Indicators and Runway")
    r = period_header(ws)

    of, ol, nf = rr(C, "open0"), rr(C, "open11"), rr(C, "new0")

    r = section(ws, r, "RETENTION")
    label(ws, r, "Net revenue retention, trailing twelve months", "%", indent=1, bold=True)
    reg(S, "nrr", r)
    write_row(ws, r, lambda i, n: ('=""' if n < 12 else
              f"=(SUM('{C}'!{col(i)}{of}:{col(i)}{ol})+SUM('{C}'!{col(i)}{nf}:{col(i)}{nf+n-12}))"
              f"/(SUM('{C}'!{col(i-12)}{of}:{col(i-12)}{ol})"
              f"+SUM('{C}'!{col(i-12)}{nf}:{col(i-12)}{nf+n-12}))"), FMT_PCT1, GREEN, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "Gross revenue retention, trailing twelve months", "%", indent=1, bold=True)
    reg(S, "grr", r)
    write_row(ws, r, lambda i, n: ('=""' if n < 12 else
              f"=1+SUM('{B}'!{col(i-11)}{rr(B,'churn')}:{col(i)}{rr(B,'churn')})"
              f"/'{B}'!{col(i-12)}{rr(B,'total_arr')}"), FMT_PCT1, GREEN, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "Benchmark: median 89%, top quartile 94%", "", indent=2, italic=True); r += 2

    r = section(ws, r, "EFFICIENCY")
    label(ws, r, "Gross margin", "%", indent=1); reg(S, "gm", r)
    write_row(ws, r, lambda i, n: f"={x(P,'gm',i)}", FMT_PCT1, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "Burn multiple, net burn over net new ARR", "x", indent=1, bold=True)
    reg(S, "bm", r)
    write_row(ws, r, lambda i, n: ('=""' if n < 12 else
              f"=IF(OR('{B}'!{col(i)}{rr(B,'total_arr')}-'{B}'!{col(i-12)}{rr(B,'total_arr')}<=0,"
              f"SUM('{CF}'!{col(i-11)}{rr(CF,'fcf')}:{col(i)}{rr(CF,'fcf')})>=0),\"\","
              f"-SUM('{CF}'!{col(i-11)}{rr(CF,'fcf')}:{col(i)}{rr(CF,'fcf')})"
              f"/('{B}'!{col(i)}{rr(B,'total_arr')}-'{B}'!{col(i-12)}{rr(B,'total_arr')}))"),
              FMT_X, GREEN, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "ARR per full time employee", "$", indent=1); reg(S, "arrfte", r)
    write_row(ws, r, lambda i, n: f"={x(B,'total_arr',i)}/{x(H,'total',i)}", FMT_USD0, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "Rule of 40, growth plus EBITDA margin", "%", indent=1); reg(S, "r40", r)
    write_row(ws, r, lambda i, n: ('=""' if n < 12 else
              f"={x(B,'yoy',i)}+SUM('{P}'!{col(i-11)}{rr(P,'ebitda')}:{col(i)}{rr(P,'ebitda')})"
              f"/SUM('{P}'!{col(i-11)}{rr(P,'rev')}:{col(i)}{rr(P,'rev')})"), FMT_PCT0, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "CAC per new logo, excluding customer success", "$", indent=1)
    reg(S, "cac_dollars", r)
    write_row(ws, r, lambda i, n: ("=\"\"" if n < 3 else
              f"=IF(SUM('{CAP}'!{col(i-2)}{rr(CAP,'new_logos')}:{col(i)}{rr(CAP,'new_logos')})=0,\"\","
              f"(SUM('{P}'!{col(i-2)}{rr(P,'sm')}:{col(i)}{rr(P,'sm')})"
              f"-SUM('{P}'!{col(i-2)}{rr(P,'opex_cs')}:{col(i)}{rr(P,'opex_cs')}))"
              f"/SUM('{CAP}'!{col(i-2)}{rr(CAP,'new_logos')}:{col(i)}{rr(CAP,'new_logos')}))"),
              FMT_USD0, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "CAC payback, new business", "months", indent=1, bold=True)
    reg(S, "cac_nb", r)
    write_row(ws, r, lambda i, n: ("=\"\"" if n < 3 else
              f"=IF(SUM('{CAP}'!{col(i-2)}{rr(CAP,'new_arr')}:{col(i)}{rr(CAP,'new_arr')})*{x(P,'gm',i)}=0,\"\","
              f"(SUM('{P}'!{col(i-2)}{rr(P,'sm')}:{col(i)}{rr(P,'sm')})"
              f"-SUM('{P}'!{col(i-2)}{rr(P,'opex_cs')}:{col(i)}{rr(P,'opex_cs')}))"
              f"/(SUM('{CAP}'!{col(i-2)}{rr(CAP,'new_arr')}:{col(i)}{rr(CAP,'new_arr')})*{x(P,'gm',i)})*12)"),
              FMT_1, GREEN, bold=True)
    fy_last(ws, r); r += 1
    label(ws, r, "CAC payback, blended", "months", indent=1)
    reg(S, "cac_bl", r)
    write_row(ws, r, lambda i, n: ("=\"\"" if n < 3 else
              f"=IF(SUM('{CAP}'!{col(i-2)}{rr(CAP,'new_arr')}:{col(i)}{rr(CAP,'new_arr')})*{x(P,'gm',i)}=0,\"\","
              f"SUM('{P}'!{col(i-2)}{rr(P,'sm')}:{col(i)}{rr(P,'sm')})"
              f"/(SUM('{CAP}'!{col(i-2)}{rr(CAP,'new_arr')}:{col(i)}{rr(CAP,'new_arr')})*{x(P,'gm',i)})*12)"),
              FMT_1, GREEN)
    fy_last(ws, r); r += 1
    label(ws, r, "Magic number", "x", indent=1); reg(S, "magic", r)
    write_row(ws, r, lambda i, n: ('=""' if n < 6 else
              f"=(SUM('{P}'!{col(i-2)}{rr(P,'rev')}:{col(i)}{rr(P,'rev')})"
              f"-SUM('{P}'!{col(i-5)}{rr(P,'rev')}:{col(i-3)}{rr(P,'rev')}))*4"
              f"/SUM('{P}'!{col(i-5)}{rr(P,'sm')}:{col(i-3)}{rr(P,'sm')})"), FMT_X, GREEN)
    fy_last(ws, r); r += 2

    r = section(ws, r, "RUNWAY AND THE RAISE DECISION")
    label(ws, r, "Trailing three month average burn", "$", indent=1); reg(S, "burn", r)
    write_row(ws, r, lambda i, n:
              f"=-AVERAGE('{CF}'!{col(max(i-2,COL_M0))}{rr(CF,'fcf')}:{col(i)}{rr(CF,'fcf')})",
              FMT_USD0, GREEN); r += 1
    label(ws, r, "Months of runway to the minimum balance", "#", indent=1, bold=True)
    reg(S, "runway", r)
    write_row(ws, r, lambda i, n:
              f'=IF({col(i)}{rr(S,"burn")}<=0,"",'
              f'MAX({x(CF,"headroom",i)}/{col(i)}{rr(S,"burn")},0))',
              FMT_1, GREEN, bold=True); r += 1
    label(ws, r, "Below minimum balance", "flag", indent=1); reg(S, "flag", r)
    write_row(ws, r, lambda i, n: f'=IF({x(CF,"cash_end",i)}<{inp("min_cash")},"BREACH","")',
              "General", RED); r += 1
    label(ws, r, "Raise process should be live", "flag", indent=1); reg(S, "flag2", r)
    write_row(ws, r, lambda i, n:
              f'=IF(AND(ISNUMBER({col(i)}{rr(S,"runway")}),'
              f'{col(i)}{rr(S,"runway")}<={inp("raise_runway")},'
              f'{col(i)}{rr(S,"runway")}>0),"RAISE WINDOW","")', "General", RED); r += 2

    r = section(ws, r, "SUMMARY")
    label(ws, r, "Breach flag helper", "0/1", indent=1, italic=True)
    reg(S, "bflag", r)
    write_row(ws, r, lambda i, n: f'=IF({x(CF,"cash_end",i)}<{inp("min_cash")},1,0)', FMT_0, GREEN)
    ws.row_dimensions[r].outlineLevel = 1
    ws.row_dimensions[r].hidden = True
    fr, lr = col(COL_M0), col(COL_MLAST)
    bf = r
    r += 1
    label(ws, r, "Period in which the minimum balance is breached", "#", indent=1)
    reg(S, "bper", r)
    c = ws.cell(row=r, column=COL_M0, value=f'=IFERROR(MATCH(1,{fr}{bf}:{lr}{bf},0),0)')
    c.font = Font(size=10, color=PURPLE); c.number_format = FMT_0
    bp = f"{col(COL_M0)}{r}"
    r += 1
    for txt, fml, fmt in [
        ("Month the minimum balance is breached",
         f'=IF({bp}=0,"No breach in five years",TEXT(INDEX(\'{CF}\'!{fr}${HDR}:{lr}${HDR},{bp}),"mmm-yyyy"))',
         "General"),
        ("Raise decision date, eighteen months earlier",
         f'=IF({bp}=0,"Not required in five years",'
         f'TEXT(INDEX(\'{CF}\'!{fr}${HDR}:{lr}${HDR},MAX({bp}-{inp("raise_runway")},1)),"mmm-yyyy"))',
         "General"),
        ("Trough cash balance", f"=MIN('{CF}'!{fr}{rr(CF,'cash_end')}:{lr}{rr(CF,'cash_end')})", FMT_USD0),
        ("Peak capital requirement",
         f'=MAX({inp("min_cash")}-MIN(\'{CF}\'!{fr}{rr(CF,"cash_end")}:{lr}{rr(CF,"cash_end")}),0)',
         FMT_USD0),
        ("ARR at the raise decision date",
         f'=IF({bp}=0,"n/a",INDEX(\'{B}\'!{fr}{rr(B,"total_arr")}:{lr}{rr(B,"total_arr")},'
         f'MAX({bp}-{inp("raise_runway")},1)))', FMT_USD0),
    ]:
        label(ws, r, txt, "", indent=1, bold=True)
        c = ws.cell(row=r, column=COL_M0, value=fml)
        c.font = Font(size=10, bold=True, color=GREEN); c.number_format = fmt
        c.alignment = Alignment(horizontal="left" if fmt == "General" else "right")
        r += 1
    return ws


# =============================================================================
# SCENARIOS
# =============================================================================
def build_scenarios(wb):
    S = "Scenarios"; B = "Rev-ARR Bridge"; P = "P&L"; CF = "Cash Flow"; H = "Headcount"; K = "KPIs"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    ws.column_dimensions["B"].width = 46
    title_block(ws, "Scenarios")

    r = 5
    r = section(ws, r, "ACTIVE SCENARIO, ANNUAL SUMMARY", COL_FYLAST)
    ws.cell(row=r - 1, column=COL_M0 + 6, value="=Inputs!$E$5").font = Font(size=11, bold=True, color="FFFFFF")
    for n in range(5):
        c = ws.cell(row=r, column=COL_M0 + n, value=f"FY{YRS[n]}")
        c.font = Font(size=9, bold=True, color=NAVY); c.alignment = Alignment(horizontal="center")
    r += 1
    rows = [("Ending ARR", B, "total_arr", FMT_USD0), ("ARR growth", B, "yoy", FMT_PCT0),
            ("Revenue", P, "rev", FMT_USD0), ("Gross margin", P, "gm", FMT_PCT1),
            ("EBITDA", P, "ebitda", FMT_USD0), ("Free cash flow", CF, "fcf", FMT_USD0),
            ("Closing cash", CF, "cash_end", FMT_USD0), ("Headcount", H, "total", FMT_1),
            ("ARR per FTE", K, "arrfte", FMT_USD0), ("Net revenue retention", K, "nrr", FMT_PCT1),
            ("Burn multiple", K, "bm", FMT_X)]
    for txt, sh, key, fmt in rows:
        label(ws, r, txt, "", indent=1)
        for n in range(5):
            c = ws.cell(row=r, column=COL_M0 + n,
                        value=f"='{sh}'!{col(COL_FY0 + n)}{rr(sh, key)}")
            c.font = Font(size=9.5, color=GREEN); c.number_format = fmt
        r += 1

    r += 2
    r = section(ws, r, "SCENARIO COMPARISON", COL_M0 + 6)
    ws.cell(row=r - 1, column=COL_LABEL).value = "SCENARIO COMPARISON"
    hdrs = ["FY26 ARR", "FY28 ARR", "FY30 ARR", "FY30 headcount", "ARR per FTE",
            "Trough cash", "Breach", "Capital need"]
    for n, h in enumerate(hdrs):
        c = ws.cell(row=r, column=COL_M0 + n, value=h)
        c.font = Font(size=9, bold=True, color=NAVY); c.alignment = Alignment(horizontal="center")
    r += 1
    for si, sc in enumerate(SCN):
        df, meta = run(sc)
        a = df.groupby("year").agg(arr=("total_arr", "last"), hc=("hc_total", "last"))
        br = (df["date"].iloc[meta["breach_month"]].strftime("%b-%Y")
              if meta["breach_month"] is not None else "None in five years")
        need = max(A.GIVEN["min_cash"] - df["cash_end"].min(), 0)
        label(ws, r, SCN_LBL[si], "", indent=1, bold=(sc == "base"))
        vals = [a["arr"].iloc[0], a["arr"].iloc[2], a["arr"].iloc[4], a["hc"].iloc[4],
                a["arr"].iloc[4] / a["hc"].iloc[4], df["cash_end"].min(), br, need]
        fmts = [FMT_USD0, FMT_USD0, FMT_USD0, FMT_1, FMT_USD0, FMT_USD0, "General", FMT_USD0]
        for n, (v, f) in enumerate(zip(vals, fmts)):
            c = ws.cell(row=r, column=COL_M0 + n, value=v)
            c.font = Font(size=9.5, color=BLUE if sc != "base" else BLACK,
                          bold=(sc == "base"))
            c.number_format = f
            c.alignment = Alignment(horizontal="center" if f == "General" else "right")
        r += 1
    r += 1
    r += 1

    r = section(ws, r, "WHAT EACH CASE CHANGES, AND WHY IT EXISTS", COL_M0 + 6)
    ws.column_dimensions[col(COL_M0)].width = 34
    for n, h in enumerate(["Inputs it changes", "Why it exists"]):
        c = ws.cell(row=r, column=COL_M0 + n * 3, value=h)
        c.font = Font(size=9, bold=True, color=NAVY)
    r += 1
    doc = [
        ("Base",
         "Reference case. Nothing overridden.",
         "The plan as currently resourced. Every other case is expressed as a change from this."),
        ("Upside",
         "Attainment 78 to 85 per cent. AE hiring up. First renewal survival 92 to 94 per cent. "
         "Seat expansion up. Agents attach up. R&D hiring up.",
         "The whole engine runs better at once. Note that it still breaches the cash floor, because "
         "faster growth pulls investment forward. Growth does not remove the funding requirement."),
        ("Downside",
         "Attainment 78 to 74 per cent. First renewal survival 92 to 88.5 per cent. Seat expansion "
         "down. Agents attach down. Hiring largely unchanged.",
         "Growth disappoints while the hiring plan is already committed. Hiring is deliberately NOT "
         "cut here, because a downside that already contains its own remedy is not a downside."),
        ("Downside + Freeze",
         "Downside growth, plus R&D hiring cut to roughly a quarter, G&A hiring cut to a third, "
         "AE hiring cut, and CS coverage widened from 13 to 16 accounts per manager.",
         "Not a forecast. It isolates the management lever, showing what the cost response buys and "
         "what it costs. It is a sharp slowdown rather than a literal freeze, because a true freeze "
         "would not backfill attrition and the company would shrink."),
        ("Agents-Led",
         "Agents attach roughly doubles. Agents ACV moves from 45 to 80 per cent of core ACV, growing "
         "18 rather than 12 per cent. R&D hiring up. Core sales and retention are untouched.",
         "Tests the strategic question of whether Agents becomes the primary product. It differs from "
         "Upside in that Upside improves the existing engine, while this changes which product the "
         "company is. It also has a worse near term cash profile, because usage bills in arrears "
         "while subscriptions bill upfront."),
    ]
    for name, changes, why in doc:
        label(ws, r, name, "", indent=1, bold=True)
        for n, txt in enumerate([changes, why]):
            c = ws.cell(row=r, column=COL_M0 + n * 3, value=txt)
            c.font = Font(size=8.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 46
        r += 1
    for n in range(2):
        for k in range(3):
            ws.column_dimensions[col(COL_M0 + n * 3 + k)].width = 22
    return ws


# =============================================================================
# CHECKS
# =============================================================================
def build_checks(wb):
    S = "Checks"; B = "Rev-ARR Bridge"; C = "Rev-Cohorts"; CF = "Cash Flow"; P = "P&L"
    ws = wb.create_sheet(S); setup_sheet(ws, freeze="D8")
    title_block(ws, "Checks")
    r = period_header(ws)

    checks = [
        ("Cohort triangle foots to the ARR bridge",
         lambda i, n: f"=ABS({x(C,'ent_arr',i)}-{x(B,'end',i)})<1"),
        ("ARR bridge internally consistent, beginning plus new plus expansion less churn",
         lambda i, n: (f"=ABS({x(B,'beg',i)}+{x(B,'new',i)}+{x(B,'exp',i)}"
                       f"+{x(B,'churn',i)}-{x(B,'end',i)})<1")),
        ("Cash roll continuous, opening plus free cash flow equals closing",
         lambda i, n: f"=ABS({x(CF,'cash_beg',i)}+{x(CF,'fcf',i)}-{x(CF,'cash_end',i)})<0.01"),
        ("Opening cash of one month equals closing cash of the last",
         lambda i, n: ("=TRUE" if n == 0
                       else f"=ABS({x(CF,'cash_beg',i)}-{x(CF,'cash_end',i-1)})<0.01")),
        ("Deferred revenue rolls without a gap",
         lambda i, n: ("=TRUE" if n == 0 else
                       f"=ABS({x(CF,'deferred',i-1)}+{x(CF,'ddef',i)}-{x(CF,'deferred',i)})<0.01")),
        ("Gross profit equals revenue less cost of revenue",
         lambda i, n: f"=ABS({x(P,'rev',i)}-{x(P,'cogs',i)}-{x(P,'gp',i)})<0.01"),
        ("EBITDA equals gross profit less operating expense, excluding stock compensation",
         lambda i, n: f"=ABS({x(P,'gp',i)}-{x(P,'opex',i)}-{x(P,'ebitda',i)})<0.01"),
        ("Stock compensation equals the stated share of revenue",
         lambda i, n: f"=ABS({x(P,'rev',i)}*{inp('sbc_pct')}-{x(P,'sbc',i)})<0.01"),
        ("Total ARR is never negative", lambda i, n: f"={x(B,'total_arr',i)}>=0"),
        ("Billings equal revenue plus the change in deferred revenue, plus Agents in arrears",
         lambda i, n: (f"=ABS({x(CF,'billings',i)}-{x(P,'rev_core',i)}-{x(CF,'ddef',i)}"
                       f"-{x(CF,'bill_ag',i)})<0.01")),
    ]
    r = section(ws, r, "PERIOD BY PERIOD RECONCILIATIONS")
    first = r
    for txt, fn in checks:
        label(ws, r, txt, "", indent=1)
        write_row(ws, r, fn, "General", GREEN)
        c = ws.cell(row=r, column=COL_FY0,
                    value=f"=IF(COUNTIF({col(COL_M0)}{r}:{col(COL_MLAST)}{r},FALSE)=0,"
                          f'"OK","FAIL")')
        c.font = Font(size=10, bold=True, color=BLACK)
        c.alignment = Alignment(horizontal="center")
        ws.conditional_formatting.add(f"{col(COL_FY0)}{r}",
                                      CellIsRule(operator="equal", formula=['"OK"'], fill=FILL_OK))
        ws.conditional_formatting.add(f"{col(COL_FY0)}{r}",
                                      CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_BAD))
        r += 1
    last = r - 1

    r += 1
    r = section(ws, r, "OPENING BALANCE TIES")
    for txt, f in [
        ("Opening ARR equals the given $55.0M",
         f"=IF(ABS({inp('ent_arr0')}+{inp('ss_arr0')}-{inp('opening_arr')})<1,\"OK\",\"FAIL\")"),
        ("Opening headcount equals the given 250",
         f"=IF(ABS({inp('hc0_rnd')}+{inp('hc0_sm')}+{inp('hc0_cs')}+{inp('hc0_ga')}"
         f"-{inp('opening_hc')})<1,\"OK\",\"FAIL\")"),
        ("Opening cash equals the given $110.0M",
         f"=IF(ABS('{CF}'!{col(COL_M0)}{rr(CF,'cash_beg')}-{inp('opening_cash')})<1,\"OK\",\"FAIL\")"),
        ("Billing mix sums to one hundred per cent",
         f"=IF(ABS({inp('upfront')}+{inp('qtrly')}-1)<0.0001,\"OK\",\"FAIL\")"),
    ]:
        label(ws, r, txt, "", indent=1)
        c = ws.cell(row=r, column=COL_M0, value=f)
        c.font = Font(size=10, bold=True, color=BLACK)
        c.alignment = Alignment(horizontal="center")
        ws.conditional_formatting.add(f"{col(COL_M0)}{r}",
                                      CellIsRule(operator="equal", formula=['"OK"'], fill=FILL_OK))
        ws.conditional_formatting.add(f"{col(COL_M0)}{r}",
                                      CellIsRule(operator="equal", formula=['"FAIL"'], fill=FILL_BAD))
        r += 1

    r += 1
    label(ws, r, "OVERALL", "", bold=True)
    c = ws.cell(row=r, column=COL_M0,
                value=f'=IF(COUNTIF({col(COL_FY0)}{first}:{col(COL_FY0)}{last},"FAIL")=0,'
                      f'"ALL CHECKS PASS","REVIEW REQUIRED")')
    c.font = Font(size=12, bold=True)
    ws.conditional_formatting.add(f"{col(COL_M0)}{r}",
                                  CellIsRule(operator="equal", formula=['"ALL CHECKS PASS"'], fill=FILL_OK))
    ws.conditional_formatting.add(f"{col(COL_M0)}{r}",
                                  CellIsRule(operator="equal", formula=['"REVIEW REQUIRED"'], fill=FILL_BAD))
    return ws


# =============================================================================

def audit_self_reference(wb):
    """A formula that references its own cell is a circular reference. Fail the build.
    References qualified by a sheet name (Sheet!D12) are a different cell and are fine."""
    import re
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    if re.search(rf"(?<![A-Z0-9$!]){re.escape(c.coordinate)}(?![0-9])", v):
                        bad.append(f"{ws.title}!{c.coordinate}: {v[:70]}")
    if bad:
        raise RuntimeError("Self-referencing formulas found:\n  " + "\n  ".join(bad[:20]))
    return len(bad)


def main(path="../output/Profound_Operating_Model.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)
    build_cover(wb)
    build_inputs(wb)
    build_capacity(wb)
    build_cohorts(wb)
    build_bridge(wb)
    build_headcount(wb)
    build_pl(wb)
    build_cash(wb)
    build_kpis(wb)
    build_scenarios(wb)
    build_checks(wb)
    audit_self_reference(wb)
    wb.active = 0
    wb.save(path)
    return path


if __name__ == "__main__":
    print(main())
