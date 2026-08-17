"""
Three-way verification across every scenario.

The same equations exist as Python, as Excel formulas, and as JavaScript. This
asserts all three agree, for all five scenarios, across all sixty months. It also
checks that the scenario switch means the same thing in each artifact, which is a
failure mode a base-case-only tie-out cannot see.

Run:  python3 verify_all.py
"""
import json
import os
import shutil
import subprocess
import sys

import numpy as np
import openpyxl

import assumptions as A
import build_excel as BX
from model import run

TOL = 0.001
XL = "../output/Profound_Operating_Model.xlsx"
TMP = "/tmp/verify_all"

ROWS = [
    ("Rev-Capacity", "prod_ae", "productive_ae"),
    ("Rev-Capacity", "new_arr", "new_arr"),
    ("Rev-ARR Bridge", "end", "ent_arr"),
    ("Rev-ARR Bridge", "ag_arr", "agents_arr"),
    ("Rev-ARR Bridge", "total_arr", "total_arr"),
    ("Headcount", "total", "hc_total"),
    ("P&L", "rev", "revenue"),
    ("P&L", "cogs", "cogs"),
    ("P&L", "ebitda", "ebitda"),
    ("P&L", "sbc", "sbc"),
    ("Cash Flow", "deferred", "deferred"),
    ("Cash Flow", "billings", "billings"),
    ("Cash Flow", "fcf", "fcf"),
    ("Cash Flow", "cash_end", "cash_end"),
    ("KPIs", "nrr", "nrr"),
    ("KPIs", "grr", "grr"),
    ("KPIs", "bm", "burn_multiple"),
    ("KPIs", "arrfte", "arr_per_fte"),
    ("KPIs", "cac_nb", "cac_payback_new_business"),
    ("KPIs", "cac_bl", "cac_payback_months"),
    ("KPIs", "cac_dollars", "cac_dollars"),
]


def hdr(t):
    print("\n" + "=" * 84)
    print(t)
    print("=" * 84)


def recalc(path, outdir):
    os.makedirs(outdir, exist_ok=True)
    subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx",
                    "--outdir", outdir, path], check=True, capture_output=True, timeout=900)
    return os.path.join(outdir, os.path.basename(path))


def xl_series(wb, sheet, key):
    r = BX.R[f"{sheet}|{key}"]
    ws = wb[sheet]
    out = []
    for c in range(BX.COL_M0, BX.COL_MLAST + 1):
        v = ws.cell(row=r, column=c).value
        out.append(v if isinstance(v, (int, float)) else None)
    return out


def compare(xv, pv, label):
    """Compare a workbook row against the engine.

    Where the engine says a metric is undefined, the workbook must also be blank.
    Skipping those cells is how a formula with an empty argument, which Excel reads
    as zero rather than blank, once slipped through this harness unnoticed.
    """
    worst, worst_at, fails = 0.0, "", 0
    for m, (a, b) in enumerate(zip(pv, xv)):
        undefined = a is None or (isinstance(a, float) and np.isnan(a))
        if undefined:
            if b is not None:
                print(f"      undefined in engine but {b!r} in workbook at {label}[{m}]")
                fails += 1
            continue
        if b is None:
            fails += 1
            continue
        err = abs(b - a) / max(abs(a), 1.0)
        if err > worst:
            worst, worst_at = err, f"{label}[{m}]"
        if err > TOL:
            fails += 1
    return worst, worst_at, fails


def main():
    BX.main(XL)
    shutil.rmtree(TMP, ignore_errors=True)

    hdr("SCENARIO SWITCH ORDER   the index must mean the same thing everywhere")
    js = json.load(open("../dashboard/assumptions.json"))
    py_order = list(A.SCENARIO_LABELS.keys())
    js_order = list(js["SCENARIO_LABELS"].keys())
    ok_order = py_order == BX.SCN == js_order
    for i, k in enumerate(py_order, 1):
        print(f"  {i}. {k:24s} python {k == py_order[i-1]!s:5s} "
              f"excel {k == BX.SCN[i-1]!s:5s} dashboard {k == js_order[i-1]!s:5s}")
    print(f"  [{'PASS' if ok_order else 'FAIL'}] all three artifacts agree on scenario order")

    hdr("EXCEL vs PYTHON   every scenario, all 60 months, tolerance 0.1%")
    total_fail = 0 if ok_order else 1
    for si, sc in enumerate(BX.SCN):
        wb_w = openpyxl.load_workbook(XL)
        wb_w["Inputs"]["D5"] = si + 1
        p = f"{TMP}/scn{si}.xlsx"
        os.makedirs(TMP, exist_ok=True)
        wb_w.save(p)
        wb = openpyxl.load_workbook(recalc(p, f"{TMP}/out{si}"), data_only=True)
        df, meta = run(sc)

        worst, worst_at, fails = 0.0, "", 0
        for sheet, key, pkey in ROWS:
            if pkey not in df:
                continue
            w, wa, f = compare(xl_series(wb, sheet, key), list(df[pkey].values), f"{sheet}|{key}")
            if w > worst:
                worst, worst_at = w, wa
            fails += f
        total_fail += fails
        print(f"  [{'PASS' if fails == 0 else 'FAIL'}] {A.SCENARIO_LABELS[sc]:20s}"
              f" worst {worst:9.5%} at {worst_at:34s} cells failing {fails}")

    hdr("JAVASCRIPT vs PYTHON")
    r = subprocess.run(["node", "tie_out.mjs"], cwd="../dashboard",
                       capture_output=True, text=True, timeout=300)
    for line in r.stdout.strip().splitlines():
        if line.startswith("  [") or line.startswith("TIE-OUT"):
            print("  " + line.strip())
    if r.returncode != 0:
        total_fail += 1

    hdr("RESULT")
    print("  ALL THREE ENGINES AGREE" if total_fail == 0
          else f"  MISMATCHES FOUND: {total_fail}")
    print("=" * 84)
    return 0 if total_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
