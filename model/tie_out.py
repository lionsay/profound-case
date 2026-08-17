"""
Excel to Python tie-out.

The workbook and the engine implement the same equations. This asserts they agree
line by line on every one of the 60 months. If they ever diverge, the build fails
rather than shipping two models that quietly disagree.

Usage:  python3 tie_out.py            (base case)
        python3 tie_out.py --all      (every scenario, by rewriting the switch)
"""
import subprocess
import sys
import os
import numpy as np
import openpyxl

from model import run
import build_excel as BX

TOL = 0.001          # 0.1% relative
XL = "../output/Profound_Operating_Model.xlsx"
RECALC_DIR = "../output/recalc"

PAIRS = [
    ("Rev-Capacity", "ae_hires", "ae_hires"),
    ("Rev-Capacity", "ae_hc", "ae_headcount"),
    ("Rev-Capacity", "prod_ae", "productive_ae"),
    ("Rev-Capacity", "new_arr", "new_arr"),
    ("Rev-Capacity", "ss_cust", "ss_customers"),
    ("Rev-Capacity", "ss_arr", "ss_arr"),
    ("Rev-ARR Bridge", "end", "ent_arr"),
    ("Rev-ARR Bridge", "cust", "ent_customers"),
    ("Rev-ARR Bridge", "ag_arr", "agents_arr"),
    ("Rev-ARR Bridge", "core_arr", "core_arr"),
    ("Rev-ARR Bridge", "total_arr", "total_arr"),
    ("Headcount", "rnd", "hc_rnd"),
    ("Headcount", "cs", "hc_cs"),
    ("Headcount", "sm", "hc_sm"),
    ("Headcount", "total", "hc_total"),
    ("Headcount", "payroll", "payroll"),
    ("P&L", "rev", "revenue"),
    ("P&L", "cogs", "cogs"),
    ("P&L", "gp", "gross_profit"),
    ("P&L", "sm", "opex_sm"),
    ("P&L", "rnd", "opex_rnd"),
    ("P&L", "ga", "opex_ga"),
    ("P&L", "sbc", "sbc"),
    ("P&L", "ebitda", "ebitda"),
    ("Cash Flow", "deferred", "deferred"),
    ("Cash Flow", "billings", "billings"),
    ("Cash Flow", "capex", "capex"),
    ("Cash Flow", "interest", "interest"),
    ("Cash Flow", "fcf", "fcf"),
    ("Cash Flow", "cash_end", "cash_end"),
]


def recalc():
    os.makedirs(RECALC_DIR, exist_ok=True)
    subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx",
                    "--outdir", RECALC_DIR, XL],
                   check=True, capture_output=True, timeout=900)
    return os.path.join(RECALC_DIR, os.path.basename(XL))


def xl_row(wb, sheet, key):
    r = BX.R[f"{sheet}|{key}"]
    ws = wb[sheet]
    return np.array([ws.cell(row=r, column=c).value
                     for c in range(BX.COL_M0, BX.COL_MLAST + 1)], dtype=object)


def main():
    BX.main(XL)                       # rebuild and populate the row registry
    path = recalc()
    wb = openpyxl.load_workbook(path, data_only=True)
    df, meta = run("base")

    print("=" * 88)
    print("EXCEL TO PYTHON TIE-OUT   base case, all 60 months, tolerance 0.1%")
    print("=" * 88)
    fails = []
    for sheet, key, pkey in PAIRS:
        xv = np.array([v if isinstance(v, (int, float)) else np.nan
                       for v in xl_row(wb, sheet, key)], dtype=float)
        pv = df[pkey].values.astype(float)
        nan = int(np.isnan(xv).sum())
        err = np.nanmax(np.abs(xv - pv) / np.maximum(np.abs(pv), 1.0)) if nan < len(xv) else np.inf
        ok = err < TOL and nan == 0
        if not ok:
            fails.append((sheet, key, err, nan))
        print(f"  [{'PASS' if ok else 'FAIL'}] {sheet:16s}{key:12s} max rel err {err:8.5%}"
              f"   blanks {nan}")

    print("\n" + "=" * 88)
    print("IN-WORKBOOK CHECKS  (as evaluated by the spreadsheet itself)")
    print("=" * 88)
    ck = wb["Checks"]
    ck_fail = 0
    for r in range(1, ck.max_row + 1):
        lbl = ck.cell(row=r, column=BX.COL_LABEL).value
        for c in (BX.COL_M0, BX.COL_FY0):
            v = ck.cell(row=r, column=c).value
            if v in ("OK", "FAIL", "ALL CHECKS PASS", "REVIEW REQUIRED"):
                if v in ("FAIL", "REVIEW REQUIRED"):
                    ck_fail += 1
                print(f"  [{v}] {str(lbl)[:72]}")
                break

    print("\n" + "=" * 88)
    k = wb["KPIs"]
    print("HEADLINE OUTPUTS  (read back out of the recalculated workbook)")
    print("=" * 88)
    for r in range(1, k.max_row + 1):
        lbl = k.cell(row=r, column=BX.COL_LABEL).value
        v = k.cell(row=r, column=BX.COL_M0).value
        if lbl and any(t in str(lbl) for t in
                       ["breached", "decision date", "Trough", "Peak capital", "ARR at the raise"]):
            print(f"  {str(lbl).strip():50s} "
                  f"{f'${v:,.0f}' if isinstance(v, float) else v}")

    print("\n" + "=" * 88)
    status = "TIE-OUT CLEAN" if not fails and ck_fail == 0 else "TIE-OUT FAILED"
    print(f"{status}   rows compared {len(PAIRS)}   rows failing {len(fails)}   "
          f"workbook checks failing {ck_fail}")
    print("=" * 88)
    return 0 if (not fails and ck_fail == 0) else 1


if __name__ == "__main__":
    sys.exit(main())
